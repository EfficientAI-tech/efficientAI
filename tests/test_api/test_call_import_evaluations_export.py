"""Integration tests for the call-import evaluation CSV export endpoint.

Focuses on the ``capture_rationale`` column-doubling behavior:

* metrics with ``capture_rationale = True`` add a sibling
  ``"<Name> - LLM Rationale"`` column populated from
  ``metric_scores[id].rationale``;
* metrics without it keep their single ``"<Name>"`` value column
  (preserving the existing export shape for old metrics / old runs).
"""

from uuid import uuid4

from app.models.database import (
    CallImport,
    CallImportEvaluation,
    CallImportEvaluationRow,
    CallImportRow,
    Metric,
    Workspace,
)
from app.models.enums import CallImportRowStatus, CallImportStatus, MetricType, MetricTrigger


def _ensure_default_workspace(db_session, org_id):
    """Tests bypass the migration, so seed the per-org Default workspace
    on demand. The first call creates it; subsequent calls reuse it."""
    ws = (
        db_session.query(Workspace)
        .filter(Workspace.organization_id == org_id, Workspace.is_default.is_(True))
        .first()
    )
    if ws is None:
        ws = Workspace(
            organization_id=org_id, name="Default", slug="default", is_default=True
        )
        db_session.add(ws)
        db_session.commit()
    return ws


def _make_call_import(db_session, org_id, *, custom_columns=None):
    workspace = _ensure_default_workspace(db_session, org_id)
    call_import = CallImport(
        id=uuid4(),
        organization_id=org_id,
        workspace_id=workspace.id,
        provider="exotel",
        original_filename="test.csv",
        column_mapping={
            "external_call_id": "CallID",
            "transcript": "Transcript",
            "recording_url": "Recording URL",
        },
        extra_columns=[],
        custom_column_mapping=custom_columns or {},
        total_rows=1,
        completed_rows=1,
        failed_rows=0,
        status=CallImportStatus.COMPLETED,
    )
    db_session.add(call_import)
    db_session.commit()
    return call_import


def _make_call_import_row(db_session, call_import, row_index=0, raw_columns=None):
    row = CallImportRow(
        id=uuid4(),
        call_import_id=call_import.id,
        organization_id=call_import.organization_id,
        row_index=row_index,
        conversation_id=f"call-{row_index}",
        recording_url="https://x/r.mp3",
        transcript="hello world",
        raw_columns=raw_columns
        or {
            "CallID": f"call-{row_index}",
            "Recording URL": "https://x/r.mp3",
            "Transcript": "hello world",
        },
        status=CallImportRowStatus.COMPLETED,
    )
    db_session.add(row)
    db_session.commit()
    return row


def _make_metric(
    db_session,
    org_id,
    *,
    name,
    capture_rationale=False,
    custom_data_type=None,
    custom_config=None,
    metric_type=MetricType.RATING.value,
):
    workspace = _ensure_default_workspace(db_session, org_id)
    metric = Metric(
        id=uuid4(),
        organization_id=org_id,
        workspace_id=workspace.id,
        name=name,
        description=f"Evaluate {name}",
        metric_type=metric_type,
        trigger=MetricTrigger.ALWAYS.value,
        enabled=True,
        is_default=False,
        custom_data_type=custom_data_type,
        custom_config=custom_config,
        capture_rationale=capture_rationale,
    )
    db_session.add(metric)
    db_session.commit()
    return metric


def _make_evaluation_with_row(
    db_session,
    *,
    call_import,
    source_row,
    metrics,
    metric_scores,
):
    evaluation = CallImportEvaluation(
        id=uuid4(),
        call_import_id=call_import.id,
        organization_id=call_import.organization_id,
        workspace_id=call_import.workspace_id,
        name="QA pass",
        selected_metric_ids=[str(m.id) for m in metrics],
        status="completed",
        total_rows=1,
        completed_rows=1,
        failed_rows=0,
    )
    db_session.add(evaluation)
    db_session.commit()

    eval_row = CallImportEvaluationRow(
        id=uuid4(),
        evaluation_id=evaluation.id,
        call_import_row_id=source_row.id,
        status="completed",
        metric_scores=metric_scores,
    )
    db_session.add(eval_row)
    db_session.commit()
    return evaluation


def _parse_csv(body: bytes) -> tuple[list[str], list[dict[str, str]]]:
    import csv as csv_module
    import io as io_module

    text = body.decode("utf-8-sig")
    reader = csv_module.DictReader(io_module.StringIO(text))
    return list(reader.fieldnames or []), list(reader)


def test_export_emits_rationale_column_for_capture_rationale_metric(
    authenticated_client, db_session, org_id, seed_org
):
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)

    enum_metric = _make_metric(
        db_session,
        org_id,
        name="Pitch Type",
        capture_rationale=True,
        custom_data_type="enum",
        custom_config={"options": ["WITH data", "WITHOUT data"]},
    )

    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[enum_metric],
        metric_scores={
            str(enum_metric.id): {
                "value": "WITH data",
                "type": "enum",
                "metric_name": "Pitch Type",
                "options": ["WITH data", "WITHOUT data"],
                "rationale": "Agent referenced 120 percent growth in turn 1.",
            }
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    headers, rows = _parse_csv(response.content)

    assert "Pitch Type" in headers
    assert "Pitch Type - LLM Rationale" in headers
    # The rationale column must come immediately after its value column so
    # the user's downloaded CSV mirrors the on-screen table.
    assert headers.index("Pitch Type - LLM Rationale") == headers.index("Pitch Type") + 1

    # Three new fixed columns surface both transcripts and the run's
    # transcript_source side-by-side. They sit before the metric columns
    # so the export reads like "context → metrics".
    assert "Production Transcript" in headers
    assert "Diarised Transcript" in headers
    assert "Evaluated Transcript Source" in headers
    assert headers.index("Production Transcript") < headers.index("Pitch Type")
    assert headers.index("Diarised Transcript") < headers.index("Pitch Type")
    assert headers.index("Evaluated Transcript Source") < headers.index("Pitch Type")

    assert len(rows) == 1
    assert rows[0]["Pitch Type"] == "WITH data"
    assert rows[0]["Pitch Type - LLM Rationale"].startswith("Agent referenced")
    assert rows[0]["Production Transcript"] == "hello world"
    # No diarisation has run on this fixture row.
    assert rows[0]["Diarised Transcript"] == ""
    # Default ``transcript_source = 'production'`` -> exported label.
    assert rows[0]["Evaluated Transcript Source"] == "Production"


def test_export_omits_rationale_column_when_capture_rationale_false(
    authenticated_client, db_session, org_id, seed_org
):
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)

    plain_metric = _make_metric(
        db_session,
        org_id,
        name="Quality",
        capture_rationale=False,
    )

    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[plain_metric],
        metric_scores={
            str(plain_metric.id): {
                "value": 0.85,
                "type": "rating",
                "metric_name": "Quality",
            }
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    headers, rows = _parse_csv(response.content)

    assert "Quality" in headers
    assert "Quality - LLM Rationale" not in headers
    assert rows[0]["Quality"] == "0.85"


def test_export_handles_missing_rationale_value_gracefully(
    authenticated_client, db_session, org_id, seed_org
):
    """If the worker stored the value but no rationale (e.g. LLM returned
    nothing for the companion key), the rationale cell should be empty,
    NOT crash the export."""
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)

    metric = _make_metric(
        db_session,
        org_id,
        name="Resolution",
        capture_rationale=True,
        custom_data_type="enum",
        custom_config={"options": ["Resolved", "Unresolved"]},
    )

    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[metric],
        metric_scores={
            str(metric.id): {
                "value": "Resolved",
                "type": "enum",
                "metric_name": "Resolution",
                "options": ["Resolved", "Unresolved"],
                # ``rationale`` intentionally omitted
            }
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    headers, rows = _parse_csv(response.content)

    assert "Resolution - LLM Rationale" in headers
    assert rows[0]["Resolution"] == "Resolved"
    assert rows[0]["Resolution - LLM Rationale"] == ""


def test_export_pairs_each_metrics_value_and_rationale_in_order(
    authenticated_client, db_session, org_id, seed_org
):
    """Mixed columns: a rationale-capturing metric AND a plain metric in the
    same export. The plain metric must NOT get a phantom rationale column,
    and the rationale column must sit immediately after its own value
    column (not somewhere else in the header row)."""
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)

    a = _make_metric(
        db_session,
        org_id,
        name="Pitch Type",
        capture_rationale=True,
        custom_data_type="enum",
        custom_config={"options": ["A", "B"]},
    )
    b = _make_metric(
        db_session,
        org_id,
        name="Connect Plus",
        capture_rationale=False,
    )

    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[a, b],
        metric_scores={
            str(a.id): {
                "value": "A",
                "type": "enum",
                "metric_name": "Pitch Type",
                "options": ["A", "B"],
                "rationale": "Reason A",
            },
            str(b.id): {
                "value": True,
                "type": "boolean",
                "metric_name": "Connect Plus",
            },
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    headers, rows = _parse_csv(response.content)

    pitch_idx = headers.index("Pitch Type")
    rationale_idx = headers.index("Pitch Type - LLM Rationale")
    connect_idx = headers.index("Connect Plus")
    assert rationale_idx == pitch_idx + 1
    # Plain metric must not get a rationale column.
    assert "Connect Plus - LLM Rationale" not in headers
    # ``Connect Plus`` is selected after the rationale-capturing metric so
    # it must appear AFTER the rationale column in the header order.
    assert connect_idx > rationale_idx
    assert rows[0]["Pitch Type"] == "A"
    assert rows[0]["Pitch Type - LLM Rationale"] == "Reason A"
    assert rows[0]["Connect Plus"] == "True"


def test_export_surfaces_both_transcripts_and_diarised_source(
    authenticated_client, db_session, org_id, seed_org
):
    """When the evaluation ran against the diarised transcript, the export
    must reflect that in the new ``Evaluated Transcript Source`` column and
    the live ``Diarised Transcript`` column must carry the worker output —
    independent of the frozen ``raw_columns`` snapshot."""
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)
    # Simulate the diarisation worker having run on this row.
    source_row.diarised_transcript = "worker output text"
    db_session.commit()

    metric = _make_metric(
        db_session, org_id, name="Quality", capture_rationale=False
    )

    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[metric],
        metric_scores={
            str(metric.id): {
                "value": 0.9,
                "type": "rating",
                "metric_name": "Quality",
            }
        },
    )
    # Flip this evaluation's transcript_source to 'diarised' to mimic the
    # user picking the diarised transcript in the Run Evaluation modal.
    evaluation.transcript_source = "diarised"
    db_session.commit()

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    headers, rows = _parse_csv(response.content)

    assert "Production Transcript" in headers
    assert "Diarised Transcript" in headers
    assert "Evaluated Transcript Source" in headers
    assert rows[0]["Production Transcript"] == "hello world"
    assert rows[0]["Diarised Transcript"] == "worker output text"
    assert rows[0]["Evaluated Transcript Source"] == "Diarised"


def test_export_flattens_multi_line_diarised_transcript(
    authenticated_client, db_session, org_id, seed_org
):
    """Diarised transcripts are stored as ``<speaker>: <text>`` lines joined
    by ``\\n`` so the in-app TranscriptView can render chat bubbles. When
    exporting to CSV those newlines explode the spreadsheet cell vertically
    — every turn becomes its own line in Excel. The export endpoint must
    collapse them to a single space-separated line so the cell stays
    readable in Excel / Google Sheets without disturbing the DB shape."""
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)
    source_row.diarised_transcript = (
        "agent: hi there\nuser: hello\nagent: how can I help"
    )
    # Production transcript with a stray newline — both transcript columns
    # are flattened so the spreadsheet stays consistent regardless of
    # which source the operator picks.
    source_row.transcript = "line one\nline two"
    db_session.commit()

    metric = _make_metric(
        db_session, org_id, name="Quality", capture_rationale=False
    )
    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[metric],
        metric_scores={
            str(metric.id): {
                "value": 1.0,
                "type": "rating",
                "metric_name": "Quality",
            }
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    _, rows = _parse_csv(response.content)

    assert (
        rows[0]["Diarised Transcript"]
        == "agent: hi there user: hello agent: how can I help"
    )
    assert "\n" not in rows[0]["Diarised Transcript"]
    assert rows[0]["Production Transcript"] == "line one line two"
    assert "\n" not in rows[0]["Production Transcript"]


def test_export_supports_xlsx_format(
    authenticated_client, db_session, org_id, seed_org
):
    """The ``format=xlsx`` query param returns a native Excel workbook
    rather than a CSV — same column layout, same flattened transcripts,
    UTF-8-native cells (no BOM dance needed)."""
    import io as io_module

    from openpyxl import load_workbook

    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)
    # Use Devanagari text + multi-line diarisation to exercise both the
    # unicode handling and the cell-flattening in the xlsx writer.
    source_row.diarised_transcript = "agent: नमस्ते\nuser: हेलो"
    db_session.commit()

    metric = _make_metric(
        db_session, org_id, name="Quality", capture_rationale=False
    )
    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[metric],
        metric_scores={
            str(metric.id): {
                "value": 0.75,
                "type": "rating",
                "metric_name": "Quality",
            }
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export",
        params={"format": "xlsx"},
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    disposition = response.headers.get("content-disposition", "")
    assert ".xlsx" in disposition

    workbook = load_workbook(io_module.BytesIO(response.content), read_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    header_row = list(next(rows_iter))
    assert "Production Transcript" in header_row
    assert "Diarised Transcript" in header_row
    assert "Quality" in header_row

    data_rows = [list(r) for r in rows_iter]
    assert len(data_rows) == 1

    diarised_idx = header_row.index("Diarised Transcript")
    quality_idx = header_row.index("Quality")
    assert data_rows[0][diarised_idx] == "agent: नमस्ते user: हेलो"
    assert data_rows[0][quality_idx] == "0.75"


def test_export_defaults_to_csv_when_format_omitted(
    authenticated_client, db_session, org_id, seed_org
):
    """Backwards-compat: callers that don't pass ``format`` still get the
    same CSV response they used to (UTF-8 BOM, ``text/csv`` content-type)."""
    call_import = _make_call_import(db_session, org_id)
    source_row = _make_call_import_row(db_session, call_import)
    metric = _make_metric(
        db_session, org_id, name="Quality", capture_rationale=False
    )
    evaluation = _make_evaluation_with_row(
        db_session,
        call_import=call_import,
        source_row=source_row,
        metrics=[metric],
        metric_scores={
            str(metric.id): {
                "value": 0.5,
                "type": "rating",
                "metric_name": "Quality",
            }
        },
    )

    response = authenticated_client.get(
        f"/api/v1/call-imports/{call_import.id}/evaluations/{evaluation.id}/export"
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("text/csv")
    disposition = response.headers.get("content-disposition", "")
    assert ".csv" in disposition
