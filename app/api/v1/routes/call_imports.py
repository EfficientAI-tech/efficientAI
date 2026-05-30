"""CSV-driven call import routes.

Users upload a CSV plus a per-batch column mapping (CSV header -> system
field). The backend persists a CallImport batch + one CallImportRow per
line, then fans the rows out to the Celery ``imports`` queue where each
row is downloaded using the telephony credential pinned on the batch.
When a row has no recording URL, the worker resolves it via the chosen
provider's call detail endpoint (Exotel today; Plivo CSVs must include
the URL).
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile, status
from loguru import logger
from sqlalchemy import desc, func, or_
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.dependencies import (
    get_api_key,
    get_organization_id,
    get_workspace_id,
    require_enterprise_feature,
)
from app.models.database import (
    CallImport,
    CallImportRow,
    CallImportSchema,
    CallImportSchemaParameter,
    CallImportTag,
    TelephonyIntegration,
)
from app.models.enums import (
    CallImportParameterType,
    CallImportRowStatus,
    CallImportStatus,
)
from app.models.schemas import (
    CallImportCancelDiarisationRequest,
    CallImportCancelDiarisationResponse,
    CallImportDetailResponse,
    CallImportDiarisationPromptDefaultResponse,
    CallImportInsightsMetric,
    CallImportInsightsResponse,
    CallImportInsightsRunPoint,
    CallImportListResponse,
    CallImportMappingUpdate,
    CallImportMetricAggregate,
    CallImportPreviewResponse,
    CallImportPreviewSheet,
    CallImportRetryFailedRowsResponse,
    CallImportResponse,
    CallImportRowIdsResponse,
    CallImportRowBulkDelete,
    CallImportRowBulkDeleteResponse,
    CallImportRowResponse,
    CallImportStartRequest,
    CallImportTranscribeRequest,
    CallImportTranscribeResponse,
    CallImportUpdate,
    CallImportUploadResponse,
)


router = APIRouter(
    prefix="/call-imports",
    tags=["Call Imports"],
    dependencies=[Depends(require_enterprise_feature("call_imports"))],
)


def _normalize_dataset(raw: Optional[str]) -> Optional[str]:
    """Trim and treat empty strings as 'no dataset' (NULL)."""
    if raw is None:
        return None
    cleaned = raw.strip()
    return cleaned or None


def _resolve_tags(
    db: Session, organization_id: UUID, tag_ids: Optional[List[UUID]]
) -> List[CallImportTag]:
    """Look up tag rows by id, scoped to the organization.

    Raises HTTPException(400) if any id is unknown for the org.
    """
    if not tag_ids:
        return []
    rows = (
        db.query(CallImportTag)
        .filter(
            CallImportTag.organization_id == organization_id,
            CallImportTag.id.in_(tag_ids),
        )
        .all()
    )
    found_ids = {row.id for row in rows}
    missing = [str(tag_id) for tag_id in tag_ids if tag_id not in found_ids]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown call_import_tag id(s): {missing}",
        )
    return rows


MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB upload cap (CSV or Excel)

# File extensions accepted by the upload + preview endpoints. Keep in
# lockstep with the frontend ``accept`` attribute on the file picker.
CSV_EXTENSIONS = (".csv",)
XLSX_EXTENSIONS = (".xlsx", ".xlsm")
ALLOWED_EXTENSIONS = CSV_EXTENSIONS + XLSX_EXTENSIONS

AUDIO_CONTENT_TYPES = {
    "wav": "audio/wav",
    "mp3": "audio/mpeg",
    "flac": "audio/flac",
    "m4a": "audio/mp4",
}


def _file_format(filename: Optional[str]) -> Optional[str]:
    """Classify ``filename`` as ``'csv'`` / ``'xlsx'`` or ``None`` if unsupported."""
    if not filename:
        return None
    name = filename.lower()
    if name.endswith(CSV_EXTENSIONS):
        return "csv"
    if name.endswith(XLSX_EXTENSIONS):
        return "xlsx"
    return None


def _audio_extension(filename: Optional[str]) -> Optional[str]:
    """Return the validated lower-case extension for a manual recording."""
    if not filename or "." not in filename:
        return None
    ext = filename.rsplit(".", 1)[-1].lower().strip()
    allowed = {fmt.lower().lstrip(".") for fmt in settings.ALLOWED_AUDIO_FORMATS}
    return ext if ext in allowed else None


def _audio_content_type(ext: str, upload_content_type: Optional[str]) -> str:
    """Prefer the browser-supplied audio content type, with a safe fallback."""
    supplied = (upload_content_type or "").strip()
    if supplied and supplied != "application/octet-stream":
        return supplied
    return AUDIO_CONTENT_TYPES.get(ext.lower(), "application/octet-stream")


def _audio_s3_key(
    organization_id: UUID, call_import_id: UUID, row_id: UUID, ext: str
) -> str:
    """Build the canonical S3 key for a manually uploaded recording."""
    from app.services.storage.s3_service import s3_service

    return (
        f"{s3_service.prefix}organizations/{organization_id}/call_imports/"
        f"{call_import_id}/{row_id}.{ext}"
    )


def _filename_stem(filename: Optional[str]) -> str:
    """Extract a cross-platform filename stem from an UploadFile name."""
    raw = (filename or "").strip()
    basename = re.split(r"[\\/]", raw)[-1] if raw else ""
    if "." in basename:
        basename = basename.rsplit(".", 1)[0]
    return basename.strip()


def _sanitize_conversation_id(raw: str) -> str:
    """Turn a filename stem into a stable conversation_id."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", raw.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._-")
    return (cleaned or "recording")[:255]


def _dedupe_conversation_id(
    base: str, counts: Dict[str, int]
) -> str:
    """Make conversation ids unique within one manual upload batch."""
    count = counts.get(base, 0) + 1
    counts[base] = count
    if count == 1:
        return base
    suffix = f"-{count}"
    return f"{base[: 255 - len(suffix)]}{suffix}"


def _normalize_header(name: str) -> str:
    return (name or "").strip().lower()


def _header_lookup(fieldnames: List[str]) -> Dict[str, str]:
    """Map normalized header -> original header for case-insensitive lookup."""
    return {_normalize_header(h): h for h in fieldnames or []}


def _resolve_mapped_header(
    mapping_value: Optional[str], header_lookup: Dict[str, str]
) -> Optional[str]:
    """Translate a user-supplied CSV header into the actual column key.

    The frontend sends headers exactly as they appear in the source file,
    but we still normalize on the server so trailing whitespace / casing
    doesn't break matching. Returns the canonical fieldname or ``None``
    if not present in the file.
    """
    if not mapping_value:
        return None
    return header_lookup.get(_normalize_header(mapping_value))


def _xlsx_cell_to_str(value: Any) -> str:
    """Coerce an openpyxl cell value to the string the rest of the
    pipeline expects.

    openpyxl returns native Python types (int, float, datetime, bool,
    None). The CSV path always works with strings, so we mirror that:
    integers stringify cleanly (no ``.0`` suffix on whole-number floats),
    datetimes use ISO-8601, booleans use SQL-style ``TRUE`` / ``FALSE``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def _coerce_parameter_value(
    raw: str,
    param_type: CallImportParameterType,
    *,
    row_idx: int,
    param_name: str,
) -> Any:
    """Validate + coerce a single CSV cell against its declared type.

    Returns the typed Python value to surface in ``raw_columns``. Empty
    strings are returned as ``None`` regardless of the parameter type so
    optional cells stay null end-to-end. Coercion failures raise a
    400 with a row-anchored message.
    """
    cell = (raw or "").strip()
    if not cell:
        return None

    if param_type == CallImportParameterType.CONVERSATION_ID:
        return cell
    if param_type == CallImportParameterType.RECORDING_URL:
        # Recording URLs are exercised by the worker (which downloads
        # them); we only do a light "starts with http" check here so a
        # paste-error surfaces immediately at upload time.
        lower = cell.lower()
        if not (lower.startswith("http://") or lower.startswith("https://")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Row {row_idx + 1}: value for '{param_name}' is not a "
                    "valid recording URL (must start with http:// or https://)."
                ),
            )
        return cell
    if param_type == CallImportParameterType.RECORDING_DATE:
        try:
            parsed_date = date.fromisoformat(cell)
        except ValueError:
            try:
                parsed_date = datetime.fromisoformat(
                    cell.replace("Z", "+00:00")
                ).date()
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Row {row_idx + 1}: value for '{param_name}' is not a "
                        f"valid recording date ({cell!r}); expected YYYY-MM-DD."
                    ),
                )
        return parsed_date.isoformat()
    if param_type == CallImportParameterType.TRANSCRIPT:
        return cell
    if param_type == CallImportParameterType.TEXT:
        return cell
    if param_type == CallImportParameterType.NUMBER:
        try:
            value = float(cell)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Row {row_idx + 1}: value for '{param_name}' is not a "
                    f"valid number ({cell!r})."
                ),
            )
        if value.is_integer():
            return int(value)
        return value
    if param_type == CallImportParameterType.BOOLEAN:
        truthy = {"true", "yes", "y", "1", "t"}
        falsy = {"false", "no", "n", "0", "f"}
        norm = cell.lower()
        if norm in truthy:
            return True
        if norm in falsy:
            return False
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Row {row_idx + 1}: value for '{param_name}' is not a "
                f"valid boolean ({cell!r})."
            ),
        )
    if param_type == CallImportParameterType.DATETIME:
        try:
            parsed = datetime.fromisoformat(cell.replace("Z", "+00:00"))
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Row {row_idx + 1}: value for '{param_name}' is not a "
                    f"valid ISO-8601 date/time ({cell!r})."
                ),
            )
        return parsed.isoformat()
    if param_type == CallImportParameterType.URL:
        lower = cell.lower()
        if not (lower.startswith("http://") or lower.startswith("https://")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Row {row_idx + 1}: value for '{param_name}' is not a "
                    "valid URL (must start with http:// or https://)."
                ),
            )
        return cell
    # Unknown types: store as text and let the next migration catch up.
    return cell


def _apply_schema_mapping(
    fieldnames: List[str],
    rows_iter: Iterable[Dict[str, str]],
    parameters: List[CallImportSchemaParameter],
    parameter_mapping: Dict[str, str],
    skipped_columns: List[str],
    *,
    source_label: str = "CSV",
    validate_only: bool = False,
) -> List[Dict[str, Any]]:
    """Schema-driven row projection: parameter -> CSV header -> typed value.

    Validates that every required schema parameter is mapped to a CSV
    header that actually exists in the file, and that every CSV header
    is either mapped to a parameter or explicitly listed in
    ``skipped_columns``. Returns one dict per non-empty data row with:

      * ``conversation_id`` (str, mandatory)
      * ``recording_date`` (Optional[str], ISO date)
      * ``recording_url`` (Optional[str])
      * ``transcript`` (Optional[str])
      * ``parameter_values`` (Dict[str, Any]) of typed values keyed by
        parameter name (drives ``raw_columns`` so the export can
        reproduce the source).

    ``validate_only=True`` runs the header / mapping / skipped-column
    checks (every check that doesn't need to read row data) and then
    returns an empty list — used by the MAP stage to validate a
    mapping payload against the cached sheet snapshot without
    re-fetching the source bytes from S3.
    """
    header_lookup = _header_lookup(list(fieldnames))

    # 1. Look up the conversation_id parameter so we can address it
    #    directly while building each row.
    conv_param = next(
        (p for p in parameters if p.type == CallImportParameterType.CONVERSATION_ID),
        None,
    )
    if conv_param is None:
        # The schema invariant should have caught this on create/update,
        # but a defensive 400 here keeps us safe against hand-rolled
        # API callers that bypassed validation.
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected schema is missing the mandatory conversation_id parameter.",
        )
    recording_date_param = next(
        (p for p in parameters if p.type == CallImportParameterType.RECORDING_DATE),
        None,
    )
    if recording_date_param is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected schema is missing the mandatory recording_date parameter.",
        )

    # 2. Resolve every mapped parameter to a canonical fieldname.
    #    Required parameters MUST resolve; optional ones may resolve to
    #    None if the user left them blank (no mapping).
    canonical_by_param: Dict[str, Optional[str]] = {}
    recording_date_param_name: Optional[str] = None
    rec_url_param_name: Optional[str] = None
    transcript_param_name: Optional[str] = None
    for param in parameters:
        mapped_header = parameter_mapping.get(param.name)
        canonical = (
            _resolve_mapped_header(mapped_header, header_lookup)
            if mapped_header
            else None
        )
        if param.is_required and canonical is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"{source_label} does not contain the column "
                    f"'{mapped_header or ''}' mapped to required parameter "
                    f"'{param.name}'."
                ),
            )
        canonical_by_param[param.name] = canonical
        if param.type == CallImportParameterType.RECORDING_DATE:
            recording_date_param_name = param.name
        elif param.type == CallImportParameterType.RECORDING_URL:
            rec_url_param_name = param.name
        elif param.type == CallImportParameterType.TRANSCRIPT:
            transcript_param_name = param.name

    # 3. Every CSV column must either be mapped to a parameter or
    #    explicitly skipped. Catches "I forgot to skip the email
    #    column" gracefully instead of dropping data silently.
    mapped_canonicals = {c for c in canonical_by_param.values() if c}
    skipped_canonicals = {
        _resolve_mapped_header(h, header_lookup)
        for h in skipped_columns
    }
    skipped_canonicals.discard(None)
    unhandled = [
        h
        for h in fieldnames
        if h not in mapped_canonicals and h not in skipped_canonicals
    ]
    if unhandled:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"{source_label} columns must either be mapped to a schema "
                f"parameter or explicitly skipped. Unhandled: {unhandled}."
            ),
        )

    conv_canonical = canonical_by_param[conv_param.name]
    rec_canonical = (
        canonical_by_param.get(rec_url_param_name)
        if rec_url_param_name
        else None
    )
    recording_date_canonical = (
        canonical_by_param.get(recording_date_param_name)
        if recording_date_param_name
        else None
    )
    transcript_canonical = (
        canonical_by_param.get(transcript_param_name)
        if transcript_param_name
        else None
    )

    if validate_only:
        # MAP-stage validation: every header check above has already
        # run; the row loop only matters at IMPORT time. Skip it (and
        # the "no data rows" guard at the bottom of the function) so
        # the caller gets a clean pass when the mapping is shaped right.
        return []

    parsed: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows_iter):
        # Drop fully-blank lines - matches the legacy parser behavior so
        # trailing-newline edge cases don't fail an otherwise-good upload.
        non_blank = any(
            (row.get(c) or "").strip()
            for c in mapped_canonicals
            if c
        )
        if not non_blank:
            continue

        conv_value = (row.get(conv_canonical) or "").strip() if conv_canonical else ""
        if not conv_value:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Row {idx + 1} is missing the '{conv_param.name}' "
                    "(conversation_id) value."
                ),
            )

        # Materialize every mapped parameter into the per-row snapshot,
        # running per-type coercion so a bad cell aborts the upload
        # rather than silently storing garbage.
        parameter_values: Dict[str, Any] = {}
        for param in parameters:
            canonical = canonical_by_param[param.name]
            if canonical is None:
                continue
            try:
                param_type = CallImportParameterType(param.type)
            except ValueError:
                param_type = CallImportParameterType.TEXT
            coerced = _coerce_parameter_value(
                row.get(canonical) or "",
                param_type,
                row_idx=idx,
                param_name=param.name,
            )
            if param.is_required and coerced is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Row {idx + 1} is missing the required "
                        f"'{param.name}' value."
                    ),
                )
            parameter_values[param.name] = coerced

        rec_value = (
            (row.get(rec_canonical) or "").strip() if rec_canonical else ""
        )
        transcript_value = (
            (row.get(transcript_canonical) or "").strip()
            if transcript_canonical
            else ""
        )
        recording_date_value = (
            parameter_values.get(recording_date_param_name)
            if recording_date_param_name
            else None
        )

        parsed.append(
            {
                "conversation_id": conv_value,
                "recording_date": recording_date_value,
                "recording_url": rec_value or None,
                "transcript": transcript_value or None,
                "parameter_values": parameter_values,
            }
        )

    if not parsed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{source_label} did not contain any data rows.",
        )

    return parsed


def _parse_csv(
    file_bytes: bytes,
    parameters: List[CallImportSchemaParameter],
    parameter_mapping: Dict[str, str],
    skipped_columns: List[str],
) -> List[Dict[str, Any]]:
    """Parse a CSV file using the resolved schema parameters."""
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded CSV is empty.",
        )

    try:
        text_stream = io.StringIO(file_bytes.decode("utf-8-sig"))
    except UnicodeDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV must be UTF-8 encoded.",
        )

    reader = csv.DictReader(text_stream)
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV is missing a header row.",
        )

    return _apply_schema_mapping(
        list(reader.fieldnames),
        reader,
        parameters,
        parameter_mapping,
        skipped_columns,
        source_label="CSV",
    )


def _open_xlsx_workbook(file_bytes: bytes):
    """Open an xlsx/xlsm workbook from in-memory bytes (read-only stream).

    Imports openpyxl lazily so the module loads even in environments that
    haven't installed the optional dep yet (e.g. lightweight tooling
    images). Surfaces a clean 400 if openpyxl is missing or the file is
    not a valid Office Open XML workbook.
    """
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded Excel file is empty.",
        )
    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                "Excel uploads require the 'openpyxl' package which is "
                "not installed in this environment."
            ),
        ) from exc

    try:
        return load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except InvalidFileException as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File is not a valid .xlsx workbook: {exc}",
        ) from exc
    except Exception as exc:  # zipfile.BadZipFile etc.
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not open Excel workbook: {exc}",
        ) from exc


def _xlsx_sheet_headers_and_rows(
    worksheet,
) -> Tuple[List[str], List[Dict[str, str]]]:
    """Read row 1 as headers and the rest as dicts of stringified cells.

    Empty trailing header cells are dropped. Duplicate headers preserve
    the first occurrence (matches ``csv.DictReader`` behavior, which
    silently drops duplicates).
    """
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []

    headers: List[str] = []
    seen: set[str] = set()
    for cell in header_row:
        name = _xlsx_cell_to_str(cell).strip()
        if not name:
            # Stop at the first blank header — treats trailing empty
            # columns as not part of the table (matches typical Excel
            # workbook conventions).
            break
        norm = name.lower()
        if norm in seen:
            continue
        seen.add(norm)
        headers.append(name)

    rows: List[Dict[str, str]] = []
    for row in iterator:
        if row is None:
            continue
        # Pad / truncate to the header length so dict construction is
        # stable even when a row has fewer / extra cells than the header.
        cells = list(row[: len(headers)])
        if len(cells) < len(headers):
            cells.extend([None] * (len(headers) - len(cells)))
        if not any(_xlsx_cell_to_str(c).strip() for c in cells):
            # Skip fully-blank rows (openpyxl read_only routinely yields
            # trailing empties when the worksheet's used range exceeds
            # the actual data).
            continue
        rows.append(
            {
                header: _xlsx_cell_to_str(value)
                for header, value in zip(headers, cells)
            }
        )

    return headers, rows


def _parse_xlsx(
    file_bytes: bytes,
    sheet_name: Optional[str],
    parameters: List[CallImportSchemaParameter],
    parameter_mapping: Dict[str, str],
    skipped_columns: List[str],
) -> List[Dict[str, Any]]:
    """Parse a single worksheet from an xlsx/xlsm workbook.

    ``sheet_name`` must match one of the workbook's sheets (case
    insensitive whitespace-trimmed match). Returns the same shape as
    :func:`_parse_csv` so the upload handler can persist either format
    through the same code path.
    """
    if not sheet_name or not sheet_name.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="sheet_name is required when uploading an Excel workbook.",
        )

    workbook = _open_xlsx_workbook(file_bytes)
    try:
        sheet_names = list(workbook.sheetnames)
        target_norm = sheet_name.strip().lower()
        match = next(
            (s for s in sheet_names if s.strip().lower() == target_norm),
            None,
        )
        if match is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Sheet '{sheet_name}' not found in workbook. "
                    f"Available sheets: {sheet_names}"
                ),
            )
        worksheet = workbook[match]
        headers, rows = _xlsx_sheet_headers_and_rows(worksheet)
    finally:
        workbook.close()

    if not headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Sheet '{sheet_name}' is missing a header row.",
        )

    return _apply_schema_mapping(
        headers,
        rows,
        parameters,
        parameter_mapping,
        skipped_columns,
        source_label=f"Sheet '{sheet_name}'",
    )


def _csv_preview_sheets(
    file_bytes: bytes, filename: Optional[str]
) -> List[CallImportPreviewSheet]:
    """Build the synthetic single-sheet preview entry for a CSV upload."""
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded CSV is empty.",
        )
    try:
        text_stream = io.StringIO(file_bytes.decode("utf-8-sig"))
    except UnicodeDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV must be UTF-8 encoded.",
        )
    reader = csv.DictReader(text_stream)
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV is missing a header row.",
        )
    headers = list(reader.fieldnames)
    row_count = 0
    for row in reader:
        # Match the parse-time skip: ignore fully blank rows so the
        # count the user sees lines up with what /upload will ingest.
        if any((v or "").strip() for v in row.values()):
            row_count += 1

    sheet_label = (filename or "sheet1").rsplit("/", 1)[-1] or "sheet1"
    return [
        CallImportPreviewSheet(
            name=sheet_label,
            headers=headers,
            row_count=row_count,
        )
    ]


def _xlsx_preview_sheets(file_bytes: bytes) -> List[CallImportPreviewSheet]:
    """List every worksheet in the workbook with its headers and row count."""
    workbook = _open_xlsx_workbook(file_bytes)
    sheets: List[CallImportPreviewSheet] = []
    try:
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            headers, rows = _xlsx_sheet_headers_and_rows(worksheet)
            sheets.append(
                CallImportPreviewSheet(
                    name=name,
                    headers=headers,
                    row_count=len(rows),
                )
            )
    finally:
        workbook.close()
    return sheets


def _parse_json_form_field(name: str, raw: Optional[str], default):
    """Decode a JSON-encoded form field with a friendly 400 on bad JSON."""
    if raw is None or raw == "":
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{name} must be valid JSON: {exc}",
        )


# ---------------------------------------------------------------------------
# Shared helpers used by the staged endpoints (UPLOAD / MAP / IMPORT) and the
# legacy one-shot ``POST /upload`` shim. Extracted here so each stage and the
# back-compat path operate on the exact same validation + persistence code.
# ---------------------------------------------------------------------------


def _source_content_type(fmt: str) -> str:
    """Return the canonical ``Content-Type`` for a parsed file format."""
    if fmt == "csv":
        return "text/csv"
    if fmt == "xlsx":
        return (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return "application/octet-stream"


def _source_s3_key(
    organization_id: UUID, call_import_id: UUID, fmt: str
) -> str:
    """Build the canonical S3 key for an upload's source file.

    Mirrors the per-row recording key convention used by
    ``process_call_import_row`` so a single prefix sweep on delete still
    cleans up both the source artefact and every fetched recording.
    """
    from app.services.storage.s3_service import s3_service

    ext = "xlsx" if fmt == "xlsx" else "csv"
    return (
        f"{s3_service.prefix}organizations/{organization_id}/call_imports/"
        f"{call_import_id}/source.{ext}"
    )


def _build_available_sheets(
    file_bytes: bytes, fmt: str, filename: Optional[str]
) -> List[CallImportPreviewSheet]:
    """Snapshot of sheets + headers cached on the batch at UPLOAD time."""
    if fmt == "csv":
        return _csv_preview_sheets(file_bytes, filename)
    return _xlsx_preview_sheets(file_bytes)


def _resolve_schema(
    db: Session,
    organization_id: UUID,
    workspace_id: UUID,
    schema_id: UUID,
) -> CallImportSchema:
    """Fetch + validate a schema row in the active workspace.

    Eager-loads ``parameters`` so callers can iterate without re-querying.
    """
    from sqlalchemy.orm import selectinload as _selectinload

    schema = (
        db.query(CallImportSchema)
        .options(_selectinload(CallImportSchema.parameters))
        .filter(
            CallImportSchema.id == schema_id,
            CallImportSchema.organization_id == organization_id,
            CallImportSchema.workspace_id == workspace_id,
        )
        .first()
    )
    if not schema:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Call import schema not found in the active workspace.",
        )
    if not list(schema.parameters):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected schema has no parameters defined.",
        )
    return schema


def _resolve_telephony_integration(
    db: Session,
    organization_id: UUID,
    telephony_integration_id: UUID,
    provider: str,
) -> TelephonyIntegration:
    """Fetch + validate a telephony credential against the requested provider."""
    integration = (
        db.query(TelephonyIntegration)
        .filter(
            TelephonyIntegration.id == telephony_integration_id,
            TelephonyIntegration.organization_id == organization_id,
        )
        .first()
    )
    if not integration:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Telephony credential not found for this organization.",
        )
    if (integration.provider or "").lower() != provider.lower():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Selected credential is for provider '{integration.provider}', "
                f"but request specified '{provider}'."
            ),
        )
    if not integration.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected telephony credential is inactive.",
        )
    return integration


def _clean_parameter_mapping(
    mapping_payload: Any,
    parameters: List[CallImportSchemaParameter],
    schema_name: str,
) -> Dict[str, str]:
    """Trim values and drop empties; reject unknown parameter names.

    Accepts an already-decoded value (dict-shaped) so the same helper
    works for the JSON-form upload path and the JSON-body PATCH path.
    """
    if not isinstance(mapping_payload, dict) or not all(
        isinstance(k, str) and (v is None or isinstance(v, str))
        for k, v in mapping_payload.items()
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "parameter_mapping must be an object of "
                "{parameter_name: csv_header}."
            ),
        )

    valid_param_names = {p.name for p in parameters}
    cleaned: Dict[str, str] = {}
    for raw_name, raw_header in mapping_payload.items():
        if raw_name not in valid_param_names:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"parameter_mapping references unknown parameter "
                    f"'{raw_name}' on schema '{schema_name}'."
                ),
            )
        header = (raw_header or "").strip()
        if header:
            cleaned[raw_name] = header
    return cleaned


def _clean_skipped_columns(skipped_payload: Any) -> List[str]:
    """Dedupe (case-insensitively) and drop blanks; preserve original casing."""
    if not isinstance(skipped_payload, list) or not all(
        isinstance(item, str) for item in skipped_payload
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="skipped_columns must be a list of header strings.",
        )
    cleaned: List[str] = []
    seen: set[str] = set()
    for item in skipped_payload:
        norm = _normalize_header(item)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        cleaned.append(item)
    return cleaned


def _parse_source_file(
    file_bytes: bytes,
    fmt: str,
    sheet_name: Optional[str],
    parameters: List[CallImportSchemaParameter],
    cleaned_mapping: Dict[str, str],
    cleaned_skipped: List[str],
) -> List[Dict[str, Any]]:
    """Run the format-appropriate parser against a buffer of file bytes."""
    if fmt == "csv":
        return _parse_csv(file_bytes, parameters, cleaned_mapping, cleaned_skipped)
    return _parse_xlsx(
        file_bytes, sheet_name, parameters, cleaned_mapping, cleaned_skipped
    )


def _materialize_rows(
    db: Session,
    call_import: CallImport,
    parsed_rows: List[Dict[str, Any]],
    organization_id: UUID,
) -> List[CallImportRow]:
    """Insert one ``CallImportRow`` per parsed row, returning the new models."""
    row_models: List[CallImportRow] = []
    for idx, row in enumerate(parsed_rows):
        # Stamp ``transcript_source='csv'`` when the upload actually
        # provided a transcript so the UI badge ("From CSV") works from
        # day one. Blank cells stay NULL so the row reads as "no
        # production transcript yet".
        csv_transcript = row["transcript"]
        row_model = CallImportRow(
            call_import_id=call_import.id,
            organization_id=organization_id,
            row_index=idx,
            conversation_id=row["conversation_id"],
            recording_date=(
                date.fromisoformat(row["recording_date"])
                if row.get("recording_date")
                else None
            ),
            recording_url=row["recording_url"],
            transcript=csv_transcript,
            transcript_source=(
                "csv" if csv_transcript and csv_transcript.strip() else None
            ),
            raw_columns=row["parameter_values"] or None,
            status=CallImportRowStatus.PENDING,
        )
        db.add(row_model)
        row_models.append(row_model)
    return row_models


def _enqueue_row_tasks(
    db: Session,
    call_import: CallImport,
    row_models: List[CallImportRow],
) -> None:
    """Fan rows out to the ``imports`` Celery queue.

    Mirrors the legacy upload handler: on enqueue failure we mark the
    individual row FAILED and keep going so the rest of the batch
    still makes progress.
    """
    from app.workers.tasks.process_call_import_row import (
        process_call_import_row_task,
    )

    for row_model in row_models:
        try:
            process_call_import_row_task.delay(str(row_model.id))
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                "Failed to enqueue call import row {} for import {}",
                row_model.id,
                call_import.id,
            )
            row_model.status = CallImportRowStatus.FAILED
            row_model.error_message = f"Failed to enqueue: {exc}"
    db.commit()


def _ensure_s3_enabled() -> None:
    """Hard-fail UPLOAD if S3 isn't configured (no local fallback)."""
    from app.services.storage.s3_service import s3_service

    if not s3_service.is_enabled():
        err = (
            s3_service.get_status_message()
            or "S3 is not enabled or not configured"
        )
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "Call uploads require S3 storage so the file can be "
                f"persisted between stages: {err}"
            ),
        )


def _validate_sheet_choice(
    fmt: str,
    sheet_name: Optional[str],
    available_sheets: Optional[List[Dict[str, Any]]],
) -> Optional[str]:
    """Normalize / validate ``sheet_name`` against the persisted snapshot.

    Returns the canonical sheet name (matching the workbook's casing)
    so downstream parsing addresses the right worksheet.
    """
    if fmt == "csv":
        if sheet_name and sheet_name.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="sheet_name is not applicable to CSV uploads.",
            )
        return None

    cleaned = (sheet_name or "").strip() or None
    if cleaned is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="sheet_name is required when the source is an Excel workbook.",
        )

    if not available_sheets:
        # Nothing to validate against (e.g. legacy batch without snapshot);
        # let downstream parsing error out instead of silently importing.
        return cleaned

    target = cleaned.strip().lower()
    for entry in available_sheets:
        name = entry.get("name") if isinstance(entry, dict) else None
        if isinstance(name, str) and name.strip().lower() == target:
            return name
    sheet_names = [
        entry.get("name") for entry in available_sheets if isinstance(entry, dict)
    ]
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            f"Sheet '{cleaned}' not found in the staged file. "
            f"Available sheets: {sheet_names}"
        ),
    )


def _tag_response_payload(tags: Optional[List[CallImportTag]]) -> List[Dict[str, Any]]:
    """Shape a CallImport's tag relationship for the upload response."""
    return [
        {
            "id": tag.id,
            "name": tag.name,
            "color": tag.color,
            "created_at": tag.created_at,
            "updated_at": tag.updated_at,
        }
        for tag in (tags or [])
    ]


@router.post(
    "/preview",
    response_model=CallImportPreviewResponse,
    operation_id="previewCallImportFile",
)
async def preview_call_import_file(
    file: UploadFile = File(...),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportPreviewResponse:
    """Inspect an uploaded CSV / Excel file and return its sheets + headers.

    Drives the column-mapping UI without forcing the frontend to parse
    CSV / xlsx itself — keeps client and server in lockstep on quoted
    fields, encodings, and Excel cell coercion. CSVs return a single
    synthetic sheet named after the filename; Excel workbooks return one
    entry per worksheet (in workbook order).
    """
    del api_key, organization_id, workspace_id, db  # auth only

    fmt = _file_format(file.filename)
    if fmt is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Unsupported file format. Allowed extensions: "
                f"{', '.join(ALLOWED_EXTENSIONS)}."
            ),
        )

    file_bytes = await file.read()
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File exceeds {MAX_UPLOAD_BYTES} bytes",
        )

    if fmt == "csv":
        sheets = _csv_preview_sheets(file_bytes, file.filename)
    else:
        sheets = _xlsx_preview_sheets(file_bytes)

    return CallImportPreviewResponse(format=fmt, sheets=sheets)


@router.post(
    "",
    response_model=CallImportResponse,
    status_code=status.HTTP_201_CREATED,
    operation_id="createCallImport",
)
async def create_call_import(
    file: UploadFile = File(
        ...,
        description="CSV / Excel file to stage. Persisted to S3 between stages.",
    ),
    dataset: str = Form(
        ...,
        description=(
            "Required free-text dataset label. Collected up-front so the "
            "batch is filterable from the moment it lands."
        ),
    ),
    tag_ids: Optional[List[UUID]] = Form(
        None,
        description="Optional list of CallImportTag ids to attach to the new batch.",
    ),
    schema_id: Optional[UUID] = Form(
        None,
        description=(
            "Optional schema pre-pick. The user can still change it during "
            "the MAP stage; provided here only so the detail page can pre-"
            "select the schema dropdown."
        ),
    ),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportResponse:
    """UPLOAD stage of the staged call-import flow.

    Persists the source file to S3 and creates a ``CallImport`` row with
    ``status='uploaded'``. No mapping, no provider, no rows yet — the
    user moves through MAP and IMPORT as separate idempotent steps.

    Dataset is collected here (rather than at IMPORT) so the batch is
    filterable from the moment it appears in the list view.
    """
    del api_key

    fmt = _file_format(file.filename)
    if fmt is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Unsupported file format. Allowed extensions: "
                f"{', '.join(ALLOWED_EXTENSIONS)}."
            ),
        )

    normalized_dataset = _normalize_dataset(dataset)
    if not normalized_dataset:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="dataset is required and must be a non-empty string.",
        )

    file_bytes = await file.read()
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File exceeds {MAX_UPLOAD_BYTES} bytes",
        )

    # Parse-now so we (a) reject garbage uploads up-front instead of
    # later in the MAP step, and (b) capture the sheets snapshot the
    # MAP UI needs without having to re-fetch the file from S3.
    sheets = _build_available_sheets(file_bytes, fmt, file.filename)

    # Optional schema pre-pick: validated only if supplied (the user is
    # allowed to set it for the first time during MAP).
    if schema_id is not None:
        _resolve_schema(db, organization_id, workspace_id, schema_id)

    tag_rows = _resolve_tags(db, organization_id, tag_ids)

    _ensure_s3_enabled()

    # Pre-generate the id so we can compute a deterministic S3 key
    # before the row is persisted, keeping ``source_s3_key`` consistent
    # with the prefix sweep used at delete-time.
    import uuid as _uuid

    call_import_id = _uuid.uuid4()
    s3_key = _source_s3_key(organization_id, call_import_id, fmt)
    content_type = _source_content_type(fmt)

    from app.services.storage.s3_service import s3_service, StorageError

    try:
        s3_service.upload_file_by_key(file_bytes, s3_key, content_type=content_type)
    except StorageError as exc:
        logger.exception(
            "Failed to upload source file to S3 for new call import {}",
            call_import_id,
        )
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Failed to persist upload to S3: {exc}",
        )

    call_import = CallImport(
        id=call_import_id,
        organization_id=organization_id,
        workspace_id=workspace_id,
        # Provider + credential aren't known until the IMPORT stage; leave
        # them NULL so the staged-vs-legacy distinction is visible at a
        # glance from the DB.
        provider=None,
        telephony_integration_id=None,
        original_filename=file.filename,
        sheet_name=None,
        dataset=normalized_dataset,
        schema_id=schema_id,
        parameter_mapping={},
        skipped_columns=[],
        column_mapping={},
        extra_columns=[],
        custom_column_mapping={},
        source_s3_key=s3_key,
        source_format=fmt,
        source_size_bytes=len(file_bytes),
        source_content_type=content_type,
        available_sheets=[sheet.model_dump() for sheet in sheets],
        total_rows=0,
        completed_rows=0,
        failed_rows=0,
        status=CallImportStatus.UPLOADED,
    )
    if tag_rows:
        call_import.tags = tag_rows

    db.add(call_import)
    try:
        db.commit()
    except Exception:
        db.rollback()
        # Best-effort cleanup of the uploaded S3 object so a failed
        # commit doesn't leak storage.
        try:
            s3_service.delete_file_by_key(s3_key)
        except Exception as cleanup_exc:  # noqa: BLE001
            logger.warning(
                "Failed to clean up orphaned S3 object {} after DB rollback: {}",
                s3_key,
                cleanup_exc,
            )
        raise

    db.refresh(call_import)
    return CallImportResponse.model_validate(call_import)


@router.patch(
    "/{call_import_id}/mapping",
    response_model=CallImportResponse,
    operation_id="updateCallImportMapping",
)
async def update_call_import_mapping(
    call_import_id: UUID,
    payload: CallImportMappingUpdate,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportResponse:
    """MAP stage of the staged call-import flow.

    Validates ``parameter_mapping`` + ``skipped_columns`` against the
    sheet headers captured at UPLOAD time and persists them on the
    batch. Idempotent: callers may submit this multiple times while
    the batch is in ``uploaded`` or ``mapped`` state.
    """
    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
            CallImport.workspace_id == workspace_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    if call_import.status not in (
        CallImportStatus.UPLOADED,
        CallImportStatus.MAPPED,
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"Cannot edit mapping on a batch in status "
                f"'{call_import.status.value}'. Mapping can only be edited "
                "before the IMPORT stage."
            ),
        )

    if not call_import.source_s3_key or not call_import.source_format:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This batch was not uploaded through the staged flow and "
                "cannot have its mapping edited."
            ),
        )

    schema = _resolve_schema(
        db, organization_id, workspace_id, payload.schema_id
    )
    parameters = list(schema.parameters)

    canonical_sheet = _validate_sheet_choice(
        call_import.source_format,
        payload.sheet_name,
        call_import.available_sheets,
    )

    # Pull the headers for the selected sheet straight out of the
    # snapshot so we don't have to re-download the file from S3 just to
    # validate the mapping.
    headers: List[str] = []
    if call_import.available_sheets:
        if canonical_sheet is None:
            # CSV: single synthetic sheet.
            entry = call_import.available_sheets[0]
            headers = list(entry.get("headers") or [])
        else:
            for entry in call_import.available_sheets:
                if not isinstance(entry, dict):
                    continue
                name = entry.get("name")
                if isinstance(name, str) and name == canonical_sheet:
                    headers = list(entry.get("headers") or [])
                    break

    cleaned_mapping = _clean_parameter_mapping(
        payload.parameter_mapping, parameters, schema.name
    )
    cleaned_skipped = _clean_skipped_columns(payload.skipped_columns)

    # Run the same per-column validation as the parse path so the user
    # gets an immediate 400 if a required parameter is left unmapped or
    # a header is neither mapped nor skipped — without needing to read
    # the file. ``validate_only`` skips the row loop (and the empty-rows
    # guard) since the row data lives in S3, not in this request.
    if headers:
        _apply_schema_mapping(
            headers,
            iter(()),
            parameters,
            cleaned_mapping,
            cleaned_skipped,
            source_label=(
                f"Sheet '{canonical_sheet}'"
                if canonical_sheet is not None
                else "CSV"
            ),
            validate_only=True,
        )

    call_import.schema_id = schema.id
    call_import.parameter_mapping = dict(cleaned_mapping)
    call_import.skipped_columns = list(cleaned_skipped)
    call_import.sheet_name = canonical_sheet
    call_import.status = CallImportStatus.MAPPED
    db.commit()
    db.refresh(call_import)
    return CallImportResponse.model_validate(call_import)


@router.post(
    "/{call_import_id}/import",
    response_model=CallImportUploadResponse,
    status_code=status.HTTP_202_ACCEPTED,
    operation_id="startCallImport",
)
async def start_call_import(
    call_import_id: UUID,
    payload: CallImportStartRequest,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportUploadResponse:
    """IMPORT stage of the staged call-import flow.

    Re-fetches the staged source file from S3, materialises one row per
    parsed data line, and fans them out to the ``imports`` Celery queue
    so the existing per-row pipeline takes over.
    """
    del api_key

    from sqlalchemy.orm import selectinload as _selectinload

    call_import = (
        db.query(CallImport)
        .options(_selectinload(CallImport.tags))
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
            CallImport.workspace_id == workspace_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    if call_import.status != CallImportStatus.MAPPED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"Cannot start import for a batch in status "
                f"'{call_import.status.value}'. Map the columns first."
            ),
        )

    if not call_import.source_s3_key or not call_import.source_format:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This batch has no staged source file and cannot be imported "
                "through the staged flow."
            ),
        )

    if not call_import.schema_id:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot start import without a mapped schema.",
        )

    schema = _resolve_schema(
        db, organization_id, workspace_id, call_import.schema_id
    )
    parameters = list(schema.parameters)

    integration = _resolve_telephony_integration(
        db,
        organization_id,
        payload.telephony_integration_id,
        payload.provider,
    )

    # Re-fetch the staged source file from S3 each time IMPORT runs so
    # the parse is always against the artefact we promised the user
    # (vs. drifting state from a half-cached buffer).
    from app.services.storage.s3_service import s3_service, StorageError

    if not s3_service.is_enabled():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "S3 is required to read the staged source file: "
                f"{s3_service.get_status_message() or 'not configured'}"
            ),
        )

    try:
        file_bytes = s3_service.download_file_by_key(call_import.source_s3_key)
    except StorageError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not read staged source file from S3: {exc}",
        )

    cleaned_skipped = _clean_skipped_columns(
        list(call_import.skipped_columns or [])
    )

    parsed_rows = _parse_source_file(
        file_bytes,
        call_import.source_format,
        call_import.sheet_name,
        parameters,
        dict(call_import.parameter_mapping or {}),
        cleaned_skipped,
    )

    call_import.provider = integration.provider
    call_import.telephony_integration_id = integration.id
    call_import.total_rows = len(parsed_rows)
    call_import.completed_rows = 0
    call_import.failed_rows = 0

    row_models = _materialize_rows(
        db, call_import, parsed_rows, organization_id
    )

    call_import.status = CallImportStatus.PROCESSING
    db.commit()
    db.refresh(call_import)

    _enqueue_row_tasks(db, call_import, row_models)

    return CallImportUploadResponse(
        id=call_import.id,
        total_rows=call_import.total_rows,
        status=call_import.status,
        dataset=call_import.dataset,
        tags=_tag_response_payload(call_import.tags),
        message=(
            f"Accepted {call_import.total_rows} rows for import. "
            "Recordings will be fetched asynchronously."
        ),
    )


@router.post(
    "/upload",
    response_model=CallImportUploadResponse,
    status_code=status.HTTP_202_ACCEPTED,
    operation_id="uploadCallImportCsv",
    deprecated=True,
)
async def upload_call_import_csv(
    file: UploadFile = File(...),
    provider: str = Form(
        ...,
        description=(
            "Telephony provider key (e.g. 'exotel', 'plivo'). Must match the "
            "selected telephony_integration_id's provider."
        ),
    ),
    telephony_integration_id: UUID = Form(
        ...,
        description=(
            "Specific TelephonyIntegration credential row to use when "
            "downloading recordings for this batch."
        ),
    ),
    schema_id: UUID = Form(
        ...,
        description=(
            "Reusable Input Parameter schema this upload is mapped against. "
            "Must belong to the active workspace."
        ),
    ),
    parameter_mapping: str = Form(
        ...,
        description=(
            "JSON-encoded ``{schema_parameter_name: source_header}`` map "
            "covering every required schema parameter. Optional parameters "
            "may be omitted or set to an empty string."
        ),
    ),
    skipped_columns: Optional[str] = Form(
        None,
        description=(
            "JSON-encoded list of source header strings the uploader has "
            "explicitly skipped. Every header in the file must either be "
            "mapped or appear here; otherwise the upload is rejected so a "
            "forgotten column never silently drops."
        ),
    ),
    dataset: Optional[str] = Form(
        None,
        description=(
            "Optional free-text dataset label for high-level segregation. "
            "Empty strings are stored as NULL."
        ),
    ),
    tag_ids: Optional[List[UUID]] = Form(
        None,
        description="Optional list of CallImportTag ids to attach to the new batch.",
    ),
    sheet_name: Optional[str] = Form(
        None,
        description=(
            "Worksheet to import when the file is an Excel workbook "
            "(.xlsx / .xlsm). REQUIRED for Excel uploads. Ignored for CSV "
            "uploads (rejected with 400 if non-empty so typos surface "
            "instead of silently importing the wrong source)."
        ),
    ),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportUploadResponse:
    """Legacy one-shot upload kept for backward compatibility.

    DEPRECATED: prefer the staged flow
    (``POST /`` → ``PATCH /{id}/mapping`` → ``POST /{id}/import``) so
    each step is idempotent and resumable. This endpoint runs all three
    stages inline in a single transaction so existing scripts /
    integrations keep working unchanged.
    """
    del api_key

    fmt = _file_format(file.filename)
    if fmt is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Unsupported file format. Allowed extensions: "
                f"{', '.join(ALLOWED_EXTENSIONS)}."
            ),
        )

    sheet_name_clean = (sheet_name or "").strip() or None
    if fmt == "csv" and sheet_name_clean is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="sheet_name is not applicable to CSV uploads.",
        )
    if fmt == "xlsx" and sheet_name_clean is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="sheet_name is required when uploading an Excel workbook.",
        )

    file_bytes = await file.read()
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File exceeds {MAX_UPLOAD_BYTES} bytes",
        )

    schema = _resolve_schema(db, organization_id, workspace_id, schema_id)
    parameters = list(schema.parameters)

    mapping_payload = _parse_json_form_field(
        "parameter_mapping", parameter_mapping, {}
    )
    cleaned_mapping = _clean_parameter_mapping(
        mapping_payload, parameters, schema.name
    )

    skipped_payload = _parse_json_form_field("skipped_columns", skipped_columns, [])
    cleaned_skipped = _clean_skipped_columns(skipped_payload)

    integration = _resolve_telephony_integration(
        db, organization_id, telephony_integration_id, provider
    )

    parsed_rows = _parse_source_file(
        file_bytes, fmt, sheet_name_clean, parameters, cleaned_mapping, cleaned_skipped
    )

    tag_rows = _resolve_tags(db, organization_id, tag_ids)

    call_import = CallImport(
        organization_id=organization_id,
        workspace_id=workspace_id,
        provider=integration.provider,
        telephony_integration_id=integration.id,
        original_filename=file.filename,
        sheet_name=sheet_name_clean,
        dataset=_normalize_dataset(dataset),
        schema_id=schema.id,
        parameter_mapping=dict(cleaned_mapping),
        skipped_columns=list(cleaned_skipped),
        # Legacy columns are left empty on new uploads; the detail page
        # falls back to ``parameter_mapping`` when ``schema_id`` is set.
        column_mapping={},
        extra_columns=[],
        custom_column_mapping={},
        total_rows=len(parsed_rows),
        completed_rows=0,
        failed_rows=0,
        status=CallImportStatus.PENDING,
    )
    if tag_rows:
        call_import.tags = tag_rows
    db.add(call_import)
    db.flush()  # populate call_import.id

    row_models = _materialize_rows(
        db, call_import, parsed_rows, organization_id
    )

    call_import.status = CallImportStatus.PROCESSING
    db.commit()
    db.refresh(call_import)

    _enqueue_row_tasks(db, call_import, row_models)

    return CallImportUploadResponse(
        id=call_import.id,
        total_rows=call_import.total_rows,
        status=call_import.status,
        dataset=call_import.dataset,
        tags=_tag_response_payload(call_import.tags),
        message=(
            f"Accepted {call_import.total_rows} rows for import. "
            "Recordings will be fetched asynchronously."
        ),
    )


@router.post(
    "/audio-upload",
    response_model=CallImportUploadResponse,
    status_code=status.HTTP_201_CREATED,
    operation_id="uploadCallImportAudio",
)
async def upload_call_import_audio(
    files: List[UploadFile] = File(
        ...,
        description="One or more manual call recording audio files.",
    ),
    dataset: str = Form(
        ...,
        description="Required free-text dataset label for the manual upload batch.",
    ),
    tag_ids: Optional[List[UUID]] = Form(
        None,
        description="Optional list of CallImportTag ids to attach to the new batch.",
    ),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportUploadResponse:
    """Persist manually uploaded recordings as completed CallImport rows.

    The rows skip the provider-download worker entirely because the audio
    bytes are already in hand. From this point onward they behave exactly
    like completed CSV-import rows: playback reads ``recording_s3_key`` and
    the existing diarisation/evaluation endpoints can operate on them.
    """

    normalized_dataset = _normalize_dataset(dataset)
    if not normalized_dataset:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="dataset is required and must be a non-empty string.",
        )
    if not files:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one audio file is required.",
        )

    _ensure_s3_enabled()
    tag_rows = _resolve_tags(db, organization_id, tag_ids)

    max_bytes = int(settings.MAX_FILE_SIZE_MB) * 1024 * 1024
    prepared: List[Dict[str, Any]] = []
    conversation_counts: Dict[str, int] = {}

    for idx, upload in enumerate(files):
        filename = upload.filename or f"recording-{idx + 1}"
        ext = _audio_extension(filename)
        if not ext:
            allowed = ", ".join(settings.ALLOWED_AUDIO_FORMATS)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported audio file '{filename}'. Allowed formats: {allowed}.",
            )

        contents = await upload.read()
        if not contents:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Audio file '{filename}' is empty.",
            )
        if len(contents) > max_bytes:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=(
                    f"Audio file '{filename}' exceeds "
                    f"{settings.MAX_FILE_SIZE_MB} MB."
                ),
            )

        base_conversation_id = _sanitize_conversation_id(_filename_stem(filename))
        conversation_id = _dedupe_conversation_id(
            base_conversation_id,
            conversation_counts,
        )
        prepared.append(
            {
                "filename": filename,
                "extension": ext,
                "content_type": _audio_content_type(ext, upload.content_type),
                "contents": contents,
                "conversation_id": conversation_id,
            }
        )

    original_filename = (
        prepared[0]["filename"]
        if len(prepared) == 1
        else f"{len(prepared)} manual recordings"
    )
    total_size = sum(len(item["contents"]) for item in prepared)
    uploaded_keys: List[str] = []

    from app.services.storage.s3_service import s3_service

    call_import = CallImport(
        organization_id=organization_id,
        workspace_id=workspace_id,
        provider=None,
        telephony_integration_id=None,
        original_filename=original_filename,
        source_format="audio",
        source_size_bytes=total_size,
        source_content_type="audio/*",
        dataset=normalized_dataset,
        total_rows=len(prepared),
        completed_rows=len(prepared),
        failed_rows=0,
        status=CallImportStatus.COMPLETED,
    )
    if tag_rows:
        call_import.tags = tag_rows

    try:
        db.add(call_import)
        db.flush()
        # The model's historical Python default is "exotel"; manual uploads
        # intentionally have no telephony provider.
        call_import.provider = None

        for idx, item in enumerate(prepared):
            row = CallImportRow(
                call_import_id=call_import.id,
                organization_id=organization_id,
                row_index=idx,
                conversation_id=item["conversation_id"],
                recording_url=None,
                transcript=None,
                transcript_source=None,
                raw_columns={"conversation_id": item["conversation_id"]},
                status=CallImportRowStatus.COMPLETED,
            )
            db.add(row)
            db.flush()

            key = _audio_s3_key(
                organization_id,
                call_import.id,
                row.id,
                item["extension"],
            )
            s3_service.upload_file_by_key(
                item["contents"],
                key,
                content_type=item["content_type"],
            )
            uploaded_keys.append(key)

            row.recording_s3_key = key
            row.recording_content_type = item["content_type"]
            row.recording_size_bytes = len(item["contents"])

        db.commit()
    except Exception as exc:
        db.rollback()
        if uploaded_keys and s3_service.is_enabled():
            try:
                s3_service.delete_keys(uploaded_keys)
            except Exception:
                logger.exception(
                    "Failed to clean up manual audio upload keys after error"
                )
        logger.exception("Failed to persist manual call recording upload")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload manual recordings: {exc}",
        ) from exc

    db.refresh(call_import)
    return CallImportUploadResponse(
        id=call_import.id,
        total_rows=call_import.total_rows,
        status=call_import.status,
        dataset=call_import.dataset,
        tags=_tag_response_payload(call_import.tags),
        message=(
            f"Uploaded {call_import.total_rows} manual recording"
            f"{'' if call_import.total_rows == 1 else 's'}."
        ),
    )


@router.get(
    "",
    response_model=CallImportListResponse,
    operation_id="listCallImports",
)
async def list_call_imports(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status_filter: Optional[CallImportStatus] = Query(None, alias="status"),
    dataset: Optional[str] = Query(
        None,
        description=(
            "Filter by exact dataset string (case-insensitive). Pass the "
            "literal value '__none__' to filter to imports with no dataset."
        ),
    ),
    tag_id: Optional[List[UUID]] = Query(
        None,
        description="Filter to imports tagged with ALL of the given tag ids.",
    ),
    source_format: Optional[str] = Query(
        None,
        description=(
            "Filter by source format. Use 'audio' for manual recordings or "
            "'__non_audio__' for CSV/Excel/legacy imports."
        ),
    ),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportListResponse:
    """List call-import batches for the active workspace, newest first.

    Scoped to (organization_id, workspace_id) so users only see imports
    for the workspace they're currently in. Supports a high-level
    ``dataset`` filter (powers the segregation dropdown at the top of
    the imports page) plus an AND-style multi-tag filter via repeated
    ``tag_id`` parameters.
    """

    query = (
        db.query(CallImport)
        .filter(
            CallImport.organization_id == organization_id,
            CallImport.workspace_id == workspace_id,
        )
    )
    if status_filter is not None:
        query = query.filter(CallImport.status == status_filter)

    source_filter = (source_format or "").strip().lower()
    if source_filter == "__non_audio__":
        query = query.filter(
            or_(CallImport.source_format.is_(None), CallImport.source_format != "audio")
        )
    elif source_filter:
        query = query.filter(func.lower(CallImport.source_format) == source_filter)

    if dataset is not None:
        if dataset == "__none__":
            query = query.filter(CallImport.dataset.is_(None))
        elif dataset.strip():
            query = query.filter(
                func.lower(CallImport.dataset) == dataset.strip().lower()
            )

    if tag_id:
        from app.models.database import CallImportTagAssignment

        for single_tag_id in tag_id:
            sub = (
                db.query(CallImportTagAssignment.call_import_id)
                .filter(CallImportTagAssignment.tag_id == single_tag_id)
                .subquery()
            )
            query = query.filter(CallImport.id.in_(sub))

    total = query.count()
    items = (
        query.order_by(desc(CallImport.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return CallImportListResponse(
        items=[CallImportResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get(
    "/datasets",
    response_model=List[str],
    operation_id="listCallImportDatasets",
)
async def list_call_import_datasets(
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> List[str]:
    """Return the distinct, non-null dataset labels in use for the active
    workspace.

    Scoped per-workspace so each workspace's Dataset dropdown only shows
    its own segregation labels.
    """
    rows = (
        db.query(CallImport.dataset)
        .filter(
            CallImport.organization_id == organization_id,
            CallImport.workspace_id == workspace_id,
            CallImport.dataset.isnot(None),
            CallImport.dataset != "",
        )
        .distinct()
        .order_by(CallImport.dataset.asc())
        .all()
    )
    return [row[0] for row in rows if row[0]]


@router.patch(
    "/{call_import_id}",
    response_model=CallImportResponse,
    operation_id="updateCallImport",
)
async def update_call_import(
    call_import_id: UUID,
    payload: CallImportUpdate,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    workspace_id: UUID = Depends(get_workspace_id),
    db: Session = Depends(get_db),
) -> CallImportResponse:
    """Edit dataset / tag assignments (and schema, pre-import) on a batch.

    ``dataset = ""`` clears the label; ``tag_ids = []`` removes all tag
    assignments. Fields omitted from the body are left untouched.

    ``schema_id`` is only honoured while the batch is in
    ``uploaded`` / ``mapped`` state — once rows have been materialised
    the schema is locked. Changing the schema resets any persisted
    mapping (the user must re-MAP) and rewinds status to ``uploaded``.
    """
    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    body = payload.model_dump(exclude_unset=True)
    if "dataset" in body:
        call_import.dataset = _normalize_dataset(body["dataset"])

    if "tag_ids" in body:
        tag_ids = body["tag_ids"] or []
        call_import.tags = _resolve_tags(db, organization_id, tag_ids)

    if "schema_id" in body and body["schema_id"] is not None:
        if call_import.status not in (
            CallImportStatus.UPLOADED,
            CallImportStatus.MAPPED,
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    f"Cannot reassign schema on a batch in status "
                    f"'{call_import.status.value}'."
                ),
            )
        new_schema = _resolve_schema(
            db, organization_id, workspace_id, body["schema_id"]
        )
        if call_import.schema_id != new_schema.id:
            # Switching schemas invalidates the persisted mapping —
            # parameter names won't line up with the new schema, so
            # reset to UPLOADED and force a fresh MAP.
            call_import.schema_id = new_schema.id
            call_import.parameter_mapping = {}
            call_import.skipped_columns = []
            call_import.sheet_name = None
            call_import.status = CallImportStatus.UPLOADED

    db.commit()
    db.refresh(call_import)
    return CallImportResponse.model_validate(call_import)


@router.get(
    "/{call_import_id}",
    response_model=CallImportDetailResponse,
    operation_id="getCallImportDetail",
)
async def get_call_import_detail(
    call_import_id: UUID,
    row_limit: int = Query(500, ge=0, le=5000),
    row_offset: int = Query(0, ge=0),
    q: Optional[str] = Query(
        None,
        description=(
            "Optional case-insensitive substring filter on "
            "``conversation_id``. When set, ``filtered_total_rows`` in "
            "the response reflects the post-filter row count so the UI "
            "can paginate against the filtered slice."
        ),
    ),
    diarised_status: Optional[str] = Query(
        None,
        description=(
            "Optional filter on ``CallImportRow.diarised_transcript_status``. "
            "Accepts one of ``pending``, ``running``, ``completed``, "
            "``failed``. When set, ``filtered_total_rows`` reflects the "
            "post-filter row count (combined with the ``q`` filter when "
            "both are supplied) so the UI can paginate against the same "
            "slice it's displaying."
        ),
        pattern="^(pending|running|completed|failed)$",
    ),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportDetailResponse:
    """Fetch a single import batch with a slice of its rows.

    ``row_limit=0`` is intentionally allowed so callers that only need the
    batch metadata (e.g. the evaluation-detail page rendering the parent's
    column mapping) can skip the rows payload entirely.
    """

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows_query = db.query(CallImportRow).filter(
        CallImportRow.call_import_id == call_import.id
    )

    search_term = (q or "").strip()
    diarised_status_filter = (diarised_status or "").strip() or None
    filtered_total_rows: Optional[int] = None
    if search_term:
        rows_query = rows_query.filter(
            CallImportRow.conversation_id.ilike(f"%{search_term}%")
        )
    if diarised_status_filter:
        rows_query = rows_query.filter(
            CallImportRow.diarised_transcript_status == diarised_status_filter
        )
    # Surface the post-filter total whenever any filter is active so
    # the UI can paginate against the slice it's actually displaying.
    if search_term or diarised_status_filter:
        filtered_total_rows = rows_query.count()

    if row_limit == 0:
        rows: List[CallImportRow] = []
    else:
        rows = (
            rows_query.order_by(CallImportRow.row_index)
            .offset(row_offset)
            .limit(row_limit)
            .all()
        )

    # Batch-wide diarisation status aggregate. One ``GROUP BY`` query
    # across the whole batch — much cheaper than paging through every
    # row to recount on the client and lets the UI render a
    # transcribe/diarise progress bar without a separate roundtrip.
    diarised_status_counts: Dict[str, int] = {}
    for status_value, count in (
        db.query(CallImportRow.diarised_transcript_status, func.count())
        .filter(CallImportRow.call_import_id == call_import.id)
        .group_by(CallImportRow.diarised_transcript_status)
        .all()
    ):
        if isinstance(status_value, str):
            diarised_status_counts[status_value] = int(count or 0)

    detail = CallImportDetailResponse.model_validate(call_import)
    detail.rows = [CallImportRowResponse.model_validate(r) for r in rows]
    detail.filtered_total_rows = filtered_total_rows
    detail.diarised_pending_rows = diarised_status_counts.get("pending", 0)
    detail.diarised_running_rows = diarised_status_counts.get("running", 0)
    detail.diarised_completed_rows = diarised_status_counts.get("completed", 0)
    detail.diarised_failed_rows = diarised_status_counts.get("failed", 0)
    return detail


@router.get(
    "/{call_import_id}/row-ids",
    response_model=CallImportRowIdsResponse,
    operation_id="listCallImportRowIds",
)
async def list_call_import_row_ids(
    call_import_id: UUID,
    q: Optional[str] = Query(
        None,
        description=(
            "Optional case-insensitive substring filter on "
            "``conversation_id``. Same semantics as the detail endpoint."
        ),
    ),
    diarised_status: Optional[str] = Query(
        None,
        description=(
            "Optional filter on ``CallImportRow.diarised_transcript_status``. "
            "Accepts ``pending`` / ``running`` / ``completed`` / ``failed``."
        ),
        pattern="^(pending|running|completed|failed)$",
    ),
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportRowIdsResponse:
    """Return every matching ``CallImportRow.id`` for cross-page bulk select.

    Lightweight companion to ``GET /{call_import_id}`` — the detail
    endpoint caps ``row_limit`` at 5000 and ships the entire row body
    on each page, so harvesting ids that way is wasteful when the
    user just wants to bulk-delete or bulk-transcribe everything that
    matches the current filters. This endpoint applies the same ``q``
    and ``diarised_status`` filters and returns only the ids.
    """
    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows_query = db.query(CallImportRow.id).filter(
        CallImportRow.call_import_id == call_import.id
    )
    search_term = (q or "").strip()
    if search_term:
        rows_query = rows_query.filter(
            CallImportRow.conversation_id.ilike(f"%{search_term}%")
        )
    status_filter = (diarised_status or "").strip() or None
    if status_filter:
        rows_query = rows_query.filter(
            CallImportRow.diarised_transcript_status == status_filter
        )

    ids = [
        row_id
        for (row_id,) in rows_query.order_by(CallImportRow.row_index).all()
    ]
    return CallImportRowIdsResponse(ids=ids, total=len(ids))


def _revoke_pending_tasks(rows: List[CallImportRow]) -> None:
    """Best-effort revoke of in-flight Celery tasks for the given rows.

    Failures are logged and swallowed — Celery's control plane is async and
    best-effort by design, and we always do an idempotent S3 cleanup
    afterwards so a missed revoke can't leak storage.
    """
    task_ids = [
        r.celery_task_id
        for r in rows
        if r.celery_task_id
        and r.status in (CallImportRowStatus.PENDING, CallImportRowStatus.PROCESSING)
    ]
    if not task_ids:
        return

    try:
        from app.workers.celery_app import celery_app

        celery_app.control.revoke(task_ids, terminate=False)
        logger.info("Revoked {} pending call-import tasks", len(task_ids))
    except Exception as exc:  # noqa: BLE001
        logger.warning("Failed to revoke pending call-import tasks: {}", exc)


def _delete_s3_objects(
    organization_id: UUID,
    call_import_id: UUID,
    rows: List[CallImportRow],
) -> tuple[int, int]:
    """Delete every recording associated with ``rows`` plus a prefix sweep.

    The prefix sweep also cleans up the staged source file written at
    UPLOAD time (``…/call_imports/{id}/source.{csv,xlsx}``) — both the
    per-row recording keys and the source artefact share the same
    organization-scoped prefix, so a single sweep covers them all.

    Returns ``(deleted_count, error_count)``. Never raises — callers proceed
    with the DB delete regardless; orphans, if any, can be cleaned up by
    re-running the same delete (it's idempotent).
    """
    from app.services.storage.s3_service import s3_service

    if not s3_service.is_enabled():
        return 0, 0

    keys = [r.recording_s3_key for r in rows if r.recording_s3_key]
    deleted = 0
    errors = 0

    if keys:
        try:
            d, errs = s3_service.delete_keys(keys)
            deleted += d
            errors += len(errs)
            if errs:
                logger.warning(
                    "S3 bulk-delete reported {} errors for call_import {}",
                    len(errs),
                    call_import_id,
                )
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                "Bulk S3 delete failed for call_import {}: {}", call_import_id, exc
            )
            errors += len(keys)

    # Belt-and-braces sweep: catch anything that landed under the import's
    # prefix but never made it into a row's recording_s3_key (narrow
    # window where the S3 upload succeeded but the DB commit didn't).
    sweep_prefix = (
        f"{s3_service.prefix}organizations/{organization_id}/"
        f"call_imports/{call_import_id}/"
    )
    try:
        d, errs = s3_service.delete_keys_by_prefix(sweep_prefix)
        deleted += d
        errors += len(errs)
    except Exception as exc:  # noqa: BLE001
        logger.exception(
            "S3 prefix sweep failed for {}: {}", sweep_prefix, exc
        )

    return deleted, errors


@router.delete(
    "/{call_import_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    operation_id="deleteCallImport",
)
async def delete_call_import(
    call_import_id: UUID,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> Response:
    """Delete a call-import batch, its rows, and every associated S3 recording.

    Idempotent and safe to retry. If the batch is still in flight we revoke
    pending Celery tasks before tearing down the rows so workers can't
    write back to deleted records.
    """

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows = (
        db.query(CallImportRow)
        .filter(CallImportRow.call_import_id == call_import.id)
        .all()
    )

    _revoke_pending_tasks(rows)

    deleted_objects, s3_errors = _delete_s3_objects(
        organization_id=organization_id,
        call_import_id=call_import.id,
        rows=rows,
    )

    db.delete(call_import)  # cascades to call_import_rows
    db.commit()

    logger.info(
        "Deleted call_import {} (org={}, rows={}, s3_objects_deleted={}, s3_errors={})",
        call_import.id,
        organization_id,
        len(rows),
        deleted_objects,
        s3_errors,
    )

    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.delete(
    "/{call_import_id}/rows/{row_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    operation_id="deleteCallImportRow",
)
async def delete_call_import_row(
    call_import_id: UUID,
    row_id: UUID,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> Response:
    """Delete a single CallImportRow and its S3 recording.

    The parent ``CallImport`` is left in place. After deletion we recompute
    its ``total_rows`` / ``completed_rows`` / ``failed_rows`` / ``status``
    so the UI's progress bar stays consistent with reality.
    """
    from app.services.storage.s3_service import s3_service

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    row = (
        db.query(CallImportRow)
        .filter(
            CallImportRow.id == row_id,
            CallImportRow.call_import_id == call_import.id,
        )
        .first()
    )
    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import row not found",
        )

    _revoke_pending_tasks([row])

    if row.recording_s3_key and s3_service.is_enabled():
        try:
            s3_service.delete_file_by_key(row.recording_s3_key)
        except Exception as exc:  # noqa: BLE001 — best-effort, DB is source of truth
            logger.warning(
                "Failed to delete S3 object {} for row {}: {}",
                row.recording_s3_key,
                row.id,
                exc,
            )

    db.delete(row)
    db.flush()

    _recompute_call_import_counters(db, call_import)
    db.commit()

    logger.info(
        "Deleted call_import_row {} (call_import={}, org={})",
        row_id,
        call_import.id,
        organization_id,
    )

    return Response(status_code=status.HTTP_204_NO_CONTENT)


def _recompute_call_import_counters(
    db: Session, call_import: CallImport
) -> None:
    """Resync ``total/completed/failed_rows`` + status on the parent batch.

    Called after row-level mutations (single delete, bulk delete) so the
    UI's progress bar stays consistent with the actual row set. The
    rules mirror :func:`delete_call_import_row` so behavior doesn't
    diverge between the per-row and bulk paths.
    """

    remaining = (
        db.query(CallImportRow)
        .filter(CallImportRow.call_import_id == call_import.id)
        .all()
    )
    completed = sum(
        1 for r in remaining if r.status == CallImportRowStatus.COMPLETED
    )
    failed = sum(
        1 for r in remaining if r.status == CallImportRowStatus.FAILED
    )
    pending_or_processing = sum(
        1
        for r in remaining
        if r.status in (CallImportRowStatus.PENDING, CallImportRowStatus.PROCESSING)
    )

    call_import.total_rows = len(remaining)
    call_import.completed_rows = completed
    call_import.failed_rows = failed
    if pending_or_processing > 0:
        call_import.status = CallImportStatus.PROCESSING
    elif len(remaining) == 0:
        # No rows left — leave the batch in its current terminal status
        # rather than synthesizing a misleading "completed". The user can
        # delete the empty batch from the UI if they want it gone.
        pass
    elif failed == 0:
        call_import.status = CallImportStatus.COMPLETED
    elif completed == 0:
        call_import.status = CallImportStatus.FAILED
    else:
        call_import.status = CallImportStatus.PARTIAL


@router.post(
    "/{call_import_id}/retry-failed",
    response_model=CallImportRetryFailedRowsResponse,
    status_code=status.HTTP_202_ACCEPTED,
    operation_id="retryFailedCallImportRows",
)
async def retry_failed_call_import_rows(
    call_import_id: UUID,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportRetryFailedRowsResponse:
    """Re-enqueue every failed import row in this batch.

    Useful when transient provider issues are resolved and the operator wants
    a one-click "try failed downloads again" pass without re-uploading the CSV.
    """
    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    failed_rows = (
        db.query(CallImportRow)
        .filter(
            CallImportRow.call_import_id == call_import.id,
            CallImportRow.status == CallImportRowStatus.FAILED,
        )
        .order_by(CallImportRow.row_index.asc())
        .all()
    )
    if not failed_rows:
        return CallImportRetryFailedRowsResponse(
            requeued=0,
            enqueue_failed=0,
            skipped=0,
        )

    from app.workers.tasks.process_call_import_row import (
        process_call_import_row_task,
    )

    # Reset rows to pending BEFORE enqueue so the UI reflects "retry in
    # progress" immediately even if the worker queue is backlogged.
    for row in failed_rows:
        row.status = CallImportRowStatus.PENDING
        row.error_message = None
        row.celery_task_id = None

    db.flush()
    _recompute_call_import_counters(db, call_import)
    db.commit()

    requeued = 0
    enqueue_failed = 0
    skipped = 0

    for row in failed_rows:
        try:
            process_call_import_row_task.delay(str(row.id))
            requeued += 1
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                "Failed to re-enqueue call import row {} for import {}",
                row.id,
                call_import.id,
            )
            db.refresh(row)
            if row.status != CallImportRowStatus.PENDING:
                skipped += 1
                continue
            row.status = CallImportRowStatus.FAILED
            row.error_message = f"Failed to enqueue retry: {exc}"
            enqueue_failed += 1

    if enqueue_failed > 0:
        db.flush()
        _recompute_call_import_counters(db, call_import)
        db.commit()

    return CallImportRetryFailedRowsResponse(
        requeued=requeued,
        enqueue_failed=enqueue_failed,
        skipped=skipped,
    )


@router.post(
    "/{call_import_id}/rows/bulk-delete",
    response_model=CallImportRowBulkDeleteResponse,
    operation_id="bulkDeleteCallImportRows",
)
async def bulk_delete_call_import_rows(
    call_import_id: UUID,
    payload: CallImportRowBulkDelete,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportRowBulkDeleteResponse:
    """Delete multiple ``CallImportRow`` rows in one request.

    Unknown / cross-tenant row ids are silently skipped — the response
    reports how many actually went away so a UI that holds onto stale
    ids (e.g. after another tab already deleted a row) doesn't 404
    the entire bulk action.
    """
    del api_key

    from app.services.storage.s3_service import s3_service

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows = (
        db.query(CallImportRow)
        .filter(
            CallImportRow.id.in_(payload.row_ids),
            CallImportRow.call_import_id == call_import.id,
        )
        .all()
    )
    if not rows:
        return CallImportRowBulkDeleteResponse(deleted=0)

    _revoke_pending_tasks(rows)

    if s3_service.is_enabled():
        for row in rows:
            if not row.recording_s3_key:
                continue
            try:
                s3_service.delete_file_by_key(row.recording_s3_key)
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "Failed to delete S3 object {} for row {}: {}",
                    row.recording_s3_key,
                    row.id,
                    exc,
                )

    deleted = 0
    for row in rows:
        db.delete(row)
        deleted += 1
    db.flush()

    _recompute_call_import_counters(db, call_import)
    db.commit()

    logger.info(
        "Bulk-deleted {} call_import_rows (call_import={}, org={})",
        deleted,
        call_import.id,
        organization_id,
    )

    return CallImportRowBulkDeleteResponse(deleted=deleted)


# ---------------------------------------------------------------------------
# Diarization / transcription endpoints
# ---------------------------------------------------------------------------


@router.get(
    "/diarisation-prompt-default",
    response_model=CallImportDiarisationPromptDefaultResponse,
    operation_id="getCallImportDiarisationPromptDefault",
)
async def get_call_import_diarisation_prompt_default(
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
) -> CallImportDiarisationPromptDefaultResponse:
    """Return the canonical LLM diariser prompt.

    The Transcribe / Run Evaluation modals call this on open so they
    can pre-fill the prompt textarea. Returning the constant from the
    backend (rather than hard-coding it in the frontend) keeps the
    fallback used by the worker and the placeholder shown in the UI
    in lock-step — operators always see the *actual* default they'd
    get if they leave the field blank.
    """
    del api_key, organization_id
    from app.workers.tasks.helpers.llm_diarisation import (
        DEFAULT_DIARIZATION_PROMPT,
    )

    return CallImportDiarisationPromptDefaultResponse(
        prompt=DEFAULT_DIARIZATION_PROMPT
    )


def _select_rows_for_transcription(
    db: Session,
    call_import: CallImport,
    payload: CallImportTranscribeRequest,
    requested_row_ids: Optional[List[UUID]] = None,
) -> tuple[List[CallImportRow], Dict[str, int]]:
    """Pick which rows to enqueue for diarisation.

    Centralises the "should this row be touched?" decision so both the
    batch endpoint and the per-row endpoint apply the same rules:

    * row must exist on the import,
    * row must have an S3 recording (otherwise nothing to transcribe),
    * if ``only_missing`` is set and ``overwrite_existing`` is not, rows
      with a non-empty ``diarised_transcript`` are skipped (the
      production ``transcript`` is intentionally ignored here — it
      lives in a separate column and is never overwritten by this
      worker).

    Returns the list of rows to enqueue plus a per-reason skip count
    that the response can surface so the user knows why a "no-op"
    happened.
    """

    query = db.query(CallImportRow).filter(
        CallImportRow.call_import_id == call_import.id
    )
    if requested_row_ids:
        query = query.filter(CallImportRow.id.in_(requested_row_ids))
    rows = query.order_by(CallImportRow.row_index.asc()).all()

    if requested_row_ids:
        found_ids = {r.id for r in rows}
        missing = [rid for rid in requested_row_ids if rid not in found_ids]
        if missing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Some row ids were not found on this import: "
                    f"{[str(m) for m in missing]}"
                ),
            )

    selected: List[CallImportRow] = []
    skip_counts: Dict[str, int] = {}

    for row in rows:
        recording = (row.recording_s3_key or "").strip()
        if not recording:
            skip_counts["no_recording"] = skip_counts.get("no_recording", 0) + 1
            continue
        existing = (row.diarised_transcript or "").strip()
        if existing and payload.only_missing and not payload.overwrite_existing:
            skip_counts["transcript_present"] = (
                skip_counts.get("transcript_present", 0) + 1
            )
            continue
        selected.append(row)

    return selected, skip_counts


@router.post(
    "/{call_import_id}/transcribe",
    response_model=CallImportTranscribeResponse,
    status_code=status.HTTP_202_ACCEPTED,
    operation_id="transcribeCallImport",
)
async def transcribe_call_import(
    call_import_id: UUID,
    payload: CallImportTranscribeRequest,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportTranscribeResponse:
    """Fan out diarization tasks for many rows in a single call.

    Returns a summary with how many rows were queued and how many were
    skipped (broken down by reason) so the UI can show a meaningful
    toast even when nothing actually got enqueued (e.g. "All 12 rows
    already have transcripts").
    """

    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows, skip_counts = _select_rows_for_transcription(
        db, call_import, payload, requested_row_ids=payload.row_ids
    )

    # Mark the about-to-be-enqueued rows as ``pending`` so the UI can
    # show a "Queued for diarisation" badge immediately, even before
    # the worker picks the row up. The worker flips it to ``running``
    # on entry.
    for row in rows:
        row.diarised_transcript_status = "pending"
        row.diarised_transcript_error = None
    db.commit()

    if not rows:
        return CallImportTranscribeResponse(
            queued=0,
            skipped_rows=sum(skip_counts.values()),
            skipped_reason_counts=skip_counts,
        )

    from app.workers.tasks.transcribe_call_import_row import (
        transcribe_call_import_row_task,
    )

    enqueued = 0
    for row in rows:
        try:
            transcribe_call_import_row_task.delay(
                str(row.id),
                payload.stt_provider,
                payload.stt_model,
                str(payload.credential_id) if payload.credential_id else None,
                payload.language,
                payload.overwrite_existing,
                None,  # run_eval_row_id — not chained from this route
                payload.diarization_llm_provider,
                payload.diarization_llm_model,
                str(payload.diarization_llm_credential_id)
                if payload.diarization_llm_credential_id
                else None,
                payload.diarization_prompt,
                payload.mode,
            )
            enqueued += 1
        except Exception as exc:
            logger.exception(
                "Failed to enqueue transcribe for row {}: {}", row.id, exc
            )
            row.diarised_transcript_status = "failed"
            row.diarised_transcript_error = f"Failed to enqueue: {exc}"
    db.commit()

    return CallImportTranscribeResponse(
        queued=enqueued,
        skipped_rows=sum(skip_counts.values()),
        skipped_reason_counts=skip_counts,
    )


@router.post(
    "/{call_import_id}/rows/{row_id}/transcribe",
    response_model=CallImportTranscribeResponse,
    status_code=status.HTTP_202_ACCEPTED,
    operation_id="transcribeCallImportRow",
)
async def transcribe_call_import_row(
    call_import_id: UUID,
    row_id: UUID,
    payload: CallImportTranscribeRequest,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportTranscribeResponse:
    """Diarize / transcribe a single row.

    Thin wrapper over the batch endpoint that hard-codes a single
    ``row_ids`` filter. Skip counts still surface so the UI can render
    "Skipped — transcript present" diagnostics consistently.
    """

    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows, skip_counts = _select_rows_for_transcription(
        db, call_import, payload, requested_row_ids=[row_id]
    )

    for row in rows:
        row.diarised_transcript_status = "pending"
        row.diarised_transcript_error = None
    db.commit()

    if not rows:
        return CallImportTranscribeResponse(
            queued=0,
            skipped_rows=sum(skip_counts.values()),
            skipped_reason_counts=skip_counts,
        )

    from app.workers.tasks.transcribe_call_import_row import (
        transcribe_call_import_row_task,
    )

    target = rows[0]
    try:
        transcribe_call_import_row_task.delay(
            str(target.id),
            payload.stt_provider,
            payload.stt_model,
            str(payload.credential_id) if payload.credential_id else None,
            payload.language,
            payload.overwrite_existing,
            None,  # run_eval_row_id — not chained from this route
            payload.diarization_llm_provider,
            payload.diarization_llm_model,
            str(payload.diarization_llm_credential_id)
            if payload.diarization_llm_credential_id
            else None,
            payload.diarization_prompt,
            payload.mode,
        )
        return CallImportTranscribeResponse(
            queued=1,
            skipped_rows=sum(skip_counts.values()),
            skipped_reason_counts=skip_counts,
        )
    except Exception as exc:
        logger.exception(
            "Failed to enqueue transcribe for row {}: {}", target.id, exc
        )
        target.diarised_transcript_status = "failed"
        target.diarised_transcript_error = f"Failed to enqueue: {exc}"
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to enqueue transcription: {exc}",
        )


# ---------------------------------------------------------------------------
# Cancel-in-flight diarisation
# ---------------------------------------------------------------------------
#
# Long-running multimodal LLM diarisation calls (especially LLM-only mode on
# slow audio) can sit in ``pending`` / ``running`` for tens of minutes when an
# upstream provider stalls. Without an abort affordance the operator's only
# recourse is to wait for Celery's ``time_limit`` to fire — which can be
# several minutes — or to manually mutate the DB. These helpers + the two
# endpoints below give the UI a first-class "Stop diarisation" button.
#
# Why ``terminate=True``: the legacy ``_revoke_pending_tasks`` helper uses
# ``terminate=False`` because it's called from delete-flow paths where the
# task may simply not get to run (a worker pulls it off the queue and drops
# it). For a user-initiated cancel we want SIGTERM to interrupt the worker
# mid-LLM call so the audio HTTP request actually aborts. ``terminate=True``
# routes SIGTERM to the executing process; ``signal="SIGTERM"`` is the
# default but we spell it out so the intent is obvious to reviewers.

# Sentinel error message stamped on cancelled rows. Read by the transcribe
# worker's finaliser (see ``app/workers/tasks/transcribe_call_import_row.py``)
# to detect a row that was cancelled mid-flight and AVOID overwriting it
# with whatever partial result the worker had managed to compute before the
# SIGTERM landed.
CANCELLED_BY_USER_ERROR: str = "Diarisation cancelled by user"


def _cancellable_diarisation_states() -> Tuple[str, ...]:
    """States that a diarisation row can be cancelled from.

    Kept as a tiny helper so adding a future ``"queued"`` / ``"retrying"``
    state only needs one edit.
    """
    return ("pending", "running")


def _revoke_diarisation_task(row: CallImportRow) -> None:
    """Best-effort revoke of a single row's diarisation Celery task.

    Always swallows control-plane exceptions — Celery's control bus is
    inherently best-effort and a missed revoke is not catastrophic
    because the DB row is already flipped to ``failed`` by the caller
    before this runs (so the UI immediately reflects the cancel; if
    the task happens to finish anyway, the worker's finaliser skips
    over the row via :data:`CANCELLED_BY_USER_ERROR`).
    """
    task_id = (row.celery_task_id or "").strip()
    if not task_id:
        return
    try:
        from app.workers.celery_app import celery_app

        celery_app.control.revoke(
            task_id, terminate=True, signal="SIGTERM"
        )
        logger.info(
            "Revoked diarisation task {} for call-import row {}",
            task_id,
            row.id,
        )
    except Exception as exc:  # noqa: BLE001 — revoke is best-effort
        logger.warning(
            "Failed to revoke diarisation task {} for row {}: {}",
            task_id,
            row.id,
            exc,
        )


def _apply_diarisation_cancel(rows: List[CallImportRow]) -> Tuple[int, int]:
    """Cancel diarisation on every cancellable row in ``rows``.

    Returns ``(cancelled, skipped)`` so the caller can build a typed
    response without re-querying the DB. The caller is responsible for
    ``db.commit()`` after this returns — we deliberately don't commit
    here so a batch endpoint can flush all rows in one transaction.
    """
    cancellable_states = _cancellable_diarisation_states()
    cancelled = 0
    skipped = 0
    for row in rows:
        if (row.diarised_transcript_status or "").lower() not in cancellable_states:
            skipped += 1
            continue
        # Flip the row state BEFORE we revoke so the UI's next poll
        # already shows the cancel, even if Celery's control plane is
        # slow to ack.
        row.diarised_transcript_status = "failed"
        row.diarised_transcript_error = CANCELLED_BY_USER_ERROR
        _revoke_diarisation_task(row)
        # Drop the task id so a follow-up retry (or a stale poll) can't
        # accidentally re-revoke or get confused.
        row.celery_task_id = None
        cancelled += 1
    return cancelled, skipped


@router.post(
    "/{call_import_id}/rows/{row_id}/cancel-diarisation",
    response_model=CallImportRowResponse,
    status_code=status.HTTP_200_OK,
    operation_id="cancelCallImportRowDiarisation",
)
async def cancel_call_import_row_diarisation(
    call_import_id: UUID,
    row_id: UUID,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportRowResponse:
    """Abort an in-flight (or queued) diarisation for a single row.

    Idempotent: calling on a row that's already terminal (``completed``
    / ``failed`` / ``idle``) returns the row unchanged with a 200, so
    the UI can fire this from a "Stop" button without having to
    pre-check the state.

    Race notes:

    * The row's ``diarised_transcript_status`` is flipped to ``failed``
      with :data:`CANCELLED_BY_USER_ERROR` BEFORE the Celery revoke,
      so the polling UI sees the cancel immediately.
    * If the worker happens to finish between our DB flip and the
      SIGTERM landing, its finaliser will detect the cancelled
      sentinel on the row and skip its own status / score writes
      (see :mod:`app.workers.tasks.transcribe_call_import_row`).
    """
    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    row = (
        db.query(CallImportRow)
        .filter(
            CallImportRow.id == row_id,
            CallImportRow.call_import_id == call_import_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import row not found",
        )

    _apply_diarisation_cancel([row])
    db.commit()
    db.refresh(row)
    return CallImportRowResponse.model_validate(row)


@router.post(
    "/{call_import_id}/cancel-diarisation",
    response_model=CallImportCancelDiarisationResponse,
    status_code=status.HTTP_200_OK,
    operation_id="cancelCallImportDiarisation",
)
async def cancel_call_import_diarisation(
    call_import_id: UUID,
    payload: Optional[CallImportCancelDiarisationRequest] = None,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportCancelDiarisationResponse:
    """Abort in-flight diarisation for many rows in a single call.

    Default body (no ``row_ids``) cancels every row in this import
    whose ``diarised_transcript_status`` is ``pending`` or
    ``running`` — the "stop everything" button. Pass ``row_ids`` to
    scope the cancel to the rows the operator has selected.

    Returns ``(cancelled, skipped)`` so the UI can render a tight
    toast ("Cancelled 3 rows · 1 skipped (already completed)").
    """
    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    base_query = db.query(CallImportRow).filter(
        CallImportRow.call_import_id == call_import_id
    )

    requested_ids = (
        payload.row_ids if payload and payload.row_ids is not None else None
    )
    if requested_ids is not None:
        if not requested_ids:
            # Empty list is "no rows requested" — treat as a no-op
            # 200 rather than 400 so the UI can pass through an empty
            # selection without a special-case.
            return CallImportCancelDiarisationResponse(cancelled=0, skipped=0)
        rows = base_query.filter(CallImportRow.id.in_(requested_ids)).all()
        found_ids = {r.id for r in rows}
        # Treat requested-but-not-found ids as ``skipped`` so the UI's
        # numbers reconcile (a stale selection that includes deleted
        # rows shouldn't 404 the whole call).
        missing = [rid for rid in requested_ids if rid not in found_ids]
        skipped_missing = len(missing)
    else:
        # Implicit "cancel every cancellable row in this import" path.
        rows = base_query.filter(
            CallImportRow.diarised_transcript_status.in_(
                list(_cancellable_diarisation_states())
            )
        ).all()
        skipped_missing = 0

    cancelled, skipped = _apply_diarisation_cancel(rows)
    db.commit()
    return CallImportCancelDiarisationResponse(
        cancelled=cancelled,
        skipped=skipped + skipped_missing,
    )


def _render_diarised_segments_text(
    segments: Optional[List[Dict[str, Any]]],
    *,
    swap: bool = False,
) -> str:
    """Render ``CallImportRow.diarised_segments`` as ``<speaker>: <text>`` lines.

    Mirrors the worker's ``_render_turns_as_text`` (kept duplicated so
    the route doesn't need to import a Celery task module just to
    rebuild the rendered transcript). Only ``agent`` and ``user`` are
    swapped — multi-party calls keep their ``speaker_N`` labels through
    a swap so we don't silently collapse a third speaker into the user
    side.
    """
    if not segments:
        return ""
    out: List[str] = []
    for turn in segments:
        if not isinstance(turn, dict):
            continue
        speaker = (turn.get("speaker") or "").strip()
        text = (turn.get("text") or "").strip()
        if not speaker or not text:
            continue
        if swap:
            if speaker == "agent":
                speaker = "user"
            elif speaker == "user":
                speaker = "agent"
        out.append(f"{speaker}: {text}")
    return "\n".join(out)


@router.post(
    "/{call_import_id}/rows/{row_id}/diarised-speaker-swap",
    response_model=CallImportRowResponse,
    operation_id="toggleCallImportRowSpeakerSwap",
)
async def toggle_call_import_row_speaker_swap(
    call_import_id: UUID,
    row_id: UUID,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportRowResponse:
    """Flip the user <-> agent mapping on a diarised row.

    The worker's "first speaker is the agent" heuristic is right most of
    the time but does fail on inbound recordings where the customer
    greets first, on recordings where the agent stays silent for the
    intro, etc. Rather than rerun the (expensive) STT + pyannote
    pipeline for those cases, we let reviewers flip the mapping in
    place: the structured ``diarised_segments`` are the source of truth
    and we re-render the plain-text ``diarised_transcript`` from them
    with the swap applied. The next CSV export will then show the
    corrected labels.

    Returns the updated row so the frontend can refresh without an
    extra round-trip.
    """

    del api_key

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    row = (
        db.query(CallImportRow)
        .filter(
            CallImportRow.id == row_id,
            CallImportRow.call_import_id == call_import_id,
            CallImportRow.organization_id == organization_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import row not found",
        )

    segments = row.diarised_segments if isinstance(row.diarised_segments, list) else None
    if not segments:
        # Without structured turns the swap toggle would have nothing to
        # re-render — surface a clear error rather than silently
        # flipping a flag the UI never read.
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This row has no structured diarised segments to swap. "
                "Re-run diarisation to generate per-speaker turns first."
            ),
        )

    new_swap = not bool(row.diarised_speaker_swap)
    row.diarised_speaker_swap = new_swap
    row.diarised_transcript = (
        _render_diarised_segments_text(segments, swap=new_swap) or None
    )
    db.commit()
    db.refresh(row)

    return CallImportRowResponse.model_validate(row)


# ---------------------------------------------------------------------------
# Cross-run insights for the import detail page
# ---------------------------------------------------------------------------


@router.get(
    "/{call_import_id}/insights",
    response_model=CallImportInsightsResponse,
    operation_id="getCallImportInsights",
)
async def get_call_import_insights(
    call_import_id: UUID,
    api_key: str = Depends(get_api_key),
    organization_id: UUID = Depends(get_organization_id),
    db: Session = Depends(get_db),
) -> CallImportInsightsResponse:
    """Aggregate signals across every evaluation run on this import.

    Powers the Insights tab on the call-import detail page: returns
    per-metric "latest run" summaries plus a trend series of mean values
    across runs so the UI can render a small line chart per metric. Also
    bundles transcript coverage stats since those are the cheapest
    pre-eval health-check (e.g. "30 of 50 rows still missing
    transcripts").
    """

    del api_key

    from app.models.database import (
        CallImportEvaluation,
        CallImportEvaluationRow,
        Metric,
    )

    call_import = (
        db.query(CallImport)
        .filter(
            CallImport.id == call_import_id,
            CallImport.organization_id == organization_id,
        )
        .first()
    )
    if not call_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Call import not found",
        )

    rows = (
        db.query(CallImportRow)
        .filter(CallImportRow.call_import_id == call_import_id)
        .all()
    )
    # A row "has a transcript" if EITHER the production (CSV) or the
    # diarised (worker) column is populated — the insights tile reports
    # the union so users see total coverage regardless of which source
    # produced the value.
    rows_with_transcript = sum(
        1
        for r in rows
        if (r.transcript or "").strip()
        or (r.diarised_transcript or "").strip()
    )
    rows_without_transcript = len(rows) - rows_with_transcript
    source_counts: Dict[str, int] = {}
    for r in rows:
        has_production = bool((r.transcript or "").strip())
        has_diarised = bool((r.diarised_transcript or "").strip())
        if has_production:
            key = r.transcript_source or "csv"
            source_counts[key] = source_counts.get(key, 0) + 1
        if has_diarised:
            source_counts["diarised"] = source_counts.get("diarised", 0) + 1

    evaluations = (
        db.query(CallImportEvaluation)
        .filter(
            CallImportEvaluation.call_import_id == call_import_id,
            CallImportEvaluation.organization_id == organization_id,
        )
        .order_by(CallImportEvaluation.created_at.asc())
        .all()
    )

    # Defer heavy lifting to the aggregation helper so this endpoint and
    # the per-run aggregate endpoint share the exact same metric
    # bucketing math (no chance of "trend" disagreeing with "latest" on
    # the same data set).
    from app.api.v1.routes.call_import_evaluations import (
        _compute_metric_aggregates,
    )

    metric_history: Dict[str, List[CallImportInsightsRunPoint]] = {}
    metric_meta: Dict[str, Metric] = {}
    metric_latest: Dict[str, CallImportMetricAggregate] = {}

    for evaluation in evaluations:
        eval_rows = (
            db.query(CallImportEvaluationRow)
            .filter(CallImportEvaluationRow.evaluation_id == evaluation.id)
            .all()
        )
        aggregates = _compute_metric_aggregates(db, evaluation, eval_rows)
        for agg in aggregates:
            if agg.metric_id not in metric_meta:
                # ``agg.metric_id`` is normally a UUID string, but the
                # aggregator also emits ids that surface in row scores
                # without a matching ``Metric`` row (e.g. a metric the
                # user deleted mid-run, or LLM-discovered slugs). Those
                # are not valid UUIDs, so coerce defensively and skip
                # the metric registry lookup when the cast fails — the
                # ``meta is None`` branch below already handles the
                # display via the values stored on ``agg`` itself.
                try:
                    metric_uuid = UUID(agg.metric_id)
                except (ValueError, AttributeError, TypeError):
                    metric_uuid = None
                if metric_uuid is not None:
                    metric_obj = (
                        db.query(Metric)
                        .filter(
                            Metric.id == metric_uuid,
                            Metric.organization_id == organization_id,
                        )
                        .first()
                    )
                    if metric_obj is not None:
                        metric_meta[agg.metric_id] = metric_obj
            history = metric_history.setdefault(agg.metric_id, [])
            history.append(
                CallImportInsightsRunPoint(
                    evaluation_id=evaluation.id,
                    name=evaluation.name,
                    created_at=evaluation.created_at,
                    mean=agg.mean,
                    completed_rows=agg.count,
                )
            )
            metric_latest[agg.metric_id] = agg

    metrics_payload: List[CallImportInsightsMetric] = []
    for metric_id, latest in metric_latest.items():
        meta = metric_meta.get(metric_id)
        metrics_payload.append(
            CallImportInsightsMetric(
                metric_id=metric_id,
                metric_name=(meta.name if meta else latest.metric_name),
                metric_type=(meta.metric_type if meta else latest.metric_type),
                latest=latest,
                trend=metric_history.get(metric_id, []),
            )
        )

    return CallImportInsightsResponse(
        call_import_id=call_import_id,
        total_rows=len(rows),
        rows_with_transcript=rows_with_transcript,
        rows_without_transcript=rows_without_transcript,
        transcript_source_counts=source_counts,
        evaluation_count=len(evaluations),
        metrics=metrics_payload,
    )
